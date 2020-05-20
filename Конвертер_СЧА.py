""" Конвертер данных СЧА и Прирост из файла, сформированного Аванкор """

import openpyxl
# import pandas as pd

import shutil
import os
# import sys

from tkinter.filedialog import askopenfilename, asksaveasfilename
# from tkinter import messagebox
from tkinter import Tk
# Tk().withdraw()
root = Tk()
root.withdraw()

from module.analiz_data import *
from module.adjustments import *
from module.functions import *


# from openpyxl.utils import get_column_letter
# from copy import copy
# from openpyxl.styles import Font

# %%

# %%

def copy_data(ws, df_avancor, rows_numbers, columns_numbers, row_begin, col_begin):
    """ Копирование данных из таблицы Аванкор в таблицу XBRL """

    for r, row in enumerate(rows_numbers):
        for c, col in enumerate(columns_numbers):
            cell_avancor = df_avancor.loc[row, col]

            check_errors(ws, cell_avancor, row_begin, r, c, col_begin)
            # if str(cell_avancor).startswith("Не установлен") \
            #         or str(cell_avancor) == 'nan':
            #     ERRORS.append(f'"{ws.title}"; '
            #                   f'строка: {row_begin + r}, '
            #                   f'колонка: {c + 1} ; '
            #                   f'параметр: "{cell_avancor}"')

            if cell_avancor != '-' and \
                    cell_avancor != "x" and \
                    cell_avancor != 0 and \
                    str(cell_avancor) != 'nan':
                ws.cell(row_begin + r, col_begin + c).value = \
                    analiz_data_all(df_avancor.loc[row, col])


# %%

def insert_id(ws, id_row, id_cols, df_id, rows_numbers, row_begin, col_begin, id_fond):
    """ Находим и вставляем идентификаторы в ячейки"""

    # словарь ВСЕХ названий идетификаторов (полные) и имен вкладок (сокращенные)
    list_id = {(df_id[k].loc[0, 0]): k for k in df_id.keys()}
    # если ошибка в этой строке, то проверить названеи вкладок в файле с идентификаторами
    # ...

    # список идентификаторов на листе
    id_on_list = [ws.cell(id_row, col).value for col in id_cols]
    # находим признаки идентификатора в строке
    for i, id_name in enumerate(id_on_list):  # полные названия идентификаторов
        # название листа в файле с идентификатрами (сокращенные названия идентификаторов)
        sheet_name_id = list_id[id_name]

        for row in range(len(rows_numbers) - 1):
            row_i = row_begin + row
            id_is = id_serch(ws, sheet_name_id, row_i, col_begin, id_fond)

            # записываем название идентификатора в таблицу xbrl
            ws.cell(row_i, id_cols[i]).value = id_is


# %%
def id_serch(ws, sheet_name_id, row_i, col_begin, id_fond):
    """ Определение идентификатора
        :ws: форма в отчетности xbrl
        :sheet_name_id: название листа с идентификатором
        :row: текущая строка в в таблице xbrl
        :col_begin: начальная колонка в таблице xbrl
        :id_fond: идентификатор фонда
        :return 'id_is': искомый идентификатор
    """

    def round_id(df_id_ws, n, id_fond=None):

        for col in range(col_begin, ws.max_column + 1):
            id_priznak = ws.cell(row_i, col).value if id_fond is None else id_fond

            # Список признаков
            df_priznak = df_id_ws[df_id_ws[n] == id_priznak]

            # Если находим признак в строке, то прерываем перебор колонок
            if id_priznak in df_priznak[n].to_list():
                break

        return df_priznak, id_priznak

    df_id_ws = df_id[sheet_name_id]
    # Меняем тип всех данных на "str"
    # (это необходимо для корректного сравнения в дальнейшем)
    df_id_ws = df_id_ws.astype(str)

    df_priznak = df_id_ws  # текущий df
    df_priznak_list = []  # список найденных df
    id_is = None  # искомый идентификатор

    n = 1
    while n < df_priznak.columns.size:
        df_priznak, id_priznak = round_id(df_priznak, n)
        # допоминаем список c 'df_priznak'
        df_priznak_list.append(df_priznak)
        if df_priznak.index.size == 1:  # одна строчка
            # идентификатор - это превый элемент в 'df_priznak'
            id_is = df_priznak[0].values[0]
            break
        elif df_priznak.index.size == 0:  # если список признаков пуст, то ищем по 'id_fond'
            if sheet_name_id == 'Биржа':  # идентификатор не найден
                id_is = df_id_ws.loc[1, 0]
            elif sheet_name_id == 'Банковский счет':
                # берем предыдущий 'df_priznak' () из 'df_priznak_list'
                # (отнимаем "2",т.к. в списке 'df_priznak_list' элементф начинаются с "0")
                df_priznak, id_priznak = round_id(df_priznak_list[n - 2], n + 1, id_fond=id_fond)
                # идентификатор - это превый элемент в 'df_priznak'
                id_is = df_priznak[0].values[0]
            else:
                id_is = 'ошибка'
                print('------>ERROR!', round_id.__name__)
                ERRORS.append(f'ERROR! - в форме "{ws.title}" не определен "{df_id_ws.loc[0, 0]}", строка:{row_i} ')

            break
        else:
            n += 1

    if id_is:
        return id_is
    else:
        id_is = 'ошибка'
        print('.......ERROR!.......')
        print('Идентификатор не найден')
        ERRORS.append(f'Идентификатор не найден: \n'
                      f'"{ws.title}", "{df_id_ws.loc[0, 0]}", строка:{row_i}')

        return id_is


# %%

def copy_id_fond_to_tbl(wb_xbrl, id_fond):
    """Записываем во все формы идентификатор фонда"""

    # Выбираем вкладки, содержащие идентификатор фонда
    # список всех вкладок
    sheet_2_with_id_fond = wb_xbrl.sheetnames
    # удаляем из списка вкладки, не содержащие идентификатор фонда
    sheet_2_with_id_fond.remove('0420502 Справка о стоимости чис')
    sheet_2_with_id_fond.remove('0420502 Справка о стоимости _56')
    sheet_2_with_id_fond.remove('0420502 Справка о стоимости _57')
    sheet_2_with_id_fond.remove('_dropDownSheet')

    # Записываем в выбранные вкладки идентификатор фонда
    for form in sheet_2_with_id_fond:
        ws = wb_xbrl[form]
        # записывает во все таблицы xbrl идентификатор фонда
        ws.cell(row=5, column=ws.max_column).value = 'Z= Идентификатор АИФ ПИФ-' + id_fond


# %%

def load_matrica():
    """загружвем данные из матрицы"""
    file_matrica = r'./Шаблоны/Матрица.xlsx'
    df_matrica = pd.read_excel(file_matrica, sheet_name='0420502', index_col=1)
    # df_matrica.head(3)
    return df_matrica


# %%

def load_avancor(file_avancor):
    # загружвем данные из отчета таблицы Аванкор

    df_avancor = pd.read_excel(file_avancor, sheet_name='TDSheet', header=None)

    # устанавливаем начальный индекс не c 0, а c 1
    df_avancor.index += 1
    df_avancor.columns += 1

    return df_avancor


# %%

def copy_period(wb_xbrl, df_avancor, df_matrica):
    """ Вставляем данные о периоде отчетности """
    # Формы:
    # 0420502 Справка о стоимости ч_2
    # 0420502 Справка о стоимости ч_3
    # 0420502 Справка о стоимости ч_4
    # 0420502 Справка о стоимости ч_5
    # 0420502 Справка о стоимости ч_6
    # 0420502 Справка о стоимости ч_7
    # 0420502 Справка о стоимости ч_8
    # 0420502 Справка о стоимости ч_9
    # 0420502 Справка о стоимости _10
    # 0420502 Справка о стоимости _11
    # 0420502 Справка о стоимости _12

    period_begin = analiz_data_data(df_avancor.loc[18, 1])
    period_end = analiz_data_data(df_avancor.loc[18, 5])

    sheet_xbrl_period = wb_xbrl.sheetnames[8:18]

    for sheet in sheet_xbrl_period:
        ws_period = wb_xbrl[sheet]
        cell_period = df_matrica.loc[sheet, 'cell_period']
        ws_period[cell_period].value = period_begin + ', ' + period_end


# %%

def check_errors(ws, cell_avancor, row_begin, r, c, col_begin):
    """ Проверка на наличие ошибок """

    # Наименование колонки в таблице xbrl
    col_name = find_parametr(ws, row_begin, col_begin + c)

    if (str(cell_avancor).startswith("Не установлен")
        or str(cell_avancor) == 'nan'
        or str(cell_avancor) == '-') \
            and col_name != 'Примечание':
        ERRORS.append(f'"{ws.title}"; '
                      f'строка({row_begin + r}), '
                      f'колонка({c + 1});\t '
                      f'параметр: "{col_name}"'
                      f' ==> "{cell_avancor}"')


def find_parametr(ws, row_begin, col):
    """ Находим название столбца при определении возможной ошибки"""

    for row in range(1, 10):
        cell = ws.cell(row_begin - row, col).value
        if str(cell).isdigit():
            row_param = row_begin - row - 1
            cell = ws.cell(row_param, col).value
            return cell


# def write_errors(ERRORS, errors_file):
#     """ Записываем ошибки в файл"""
#     if not ERRORS:
#         ERRORS.append('Ошибок не выявлено!')
#     with open(errors_file, "w") as file:
#         for k in ERRORS:
#             file.write(str(k) + '\n\n')


# %%

def file_open(df_id):
    """ Выбор файла, созданного в Аванкор"""

    # Список всех идентификаторой фондов
    all_id_fond = df_id['ПИФ'][0][1:].to_list()

    # Выбираем файл, сформированный Аванкор
    print(f'Выбираем файл, сформированный Аванкор....'
          f'(файл должен начинаться с идентификатора фонда)')
    # show an "Open" dialog box and return the path to the selected file
    file_open = askopenfilename(initialdir="./#Отчетность",
                                title="Выбираем файл, сформированный Аванкор....",
                                filetypes=(("xlsx files", "*.xlsx"), ("All files", "*.*")))
    # Имя файла без пути к нему
    file_avancor = os.path.basename(file_open)
    # Имя файла без расширения (идентификатор фонда)
    id_fond = os.path.splitext(file_avancor)[0]

    id_fond = id_fond.split('_', 2)
    id_fond = '_'.join(id_fond[:2])

    # Если в названии файла нет идентификатора, то прерываем программу
    if not (id_fond in all_id_fond):
        print('.......ERROR!.......')
        ERRORS.append(f'{"=" * 100}\n'
                      f'Файл не сформирован!\n'
                      f'В названии файла: "{file_avancor}" неверно указан идентификатор фонда!\n'
                      f'(проверьте название файла)\n'
                      f'{"=" * 100}')
        write_errors(ERRORS, errors_file)
        sys.exit("Ошибка в имени файла!")

    print(f'выбран файл: {file_avancor}')
    return file_open, id_fond


# %%

def insert_00(ws):
    """ Проставляем нулевые значения в форме '0420502 Справка о стоимости ч_3' """

    def perebor(cell: str, col: int):
        """ Перебор колонок и проставление нулей """
        if cell:  # есть данные
            if not ws.cell(row, col + 2).value:  # в колонке с 'долей' нет данных
                ws.cell(row, col + 2).value = '0.00'
        else:
            ws.cell(row, col).value = '0.00'
            ws.cell(row, col + 2).value = '0.00'

    row_begin = 10
    col_1 = 3  # номер первой колонки с данными

    for row in range(row_begin, ws.max_row + 1):
        cell_1 = ws.cell(row, col_1).value
        cell_2 = ws.cell(row, col_1 + 1).value
        cell_3 = ws.cell(row, col_1 + 2).value
        cell_4 = ws.cell(row, col_1 + 3).value

        if cell_1 or cell_2:  # Если есть данные в одной из ячеек
            if not cell_1:
                ws.cell(row, col_1).value = '0.00'
            if not cell_2:
                ws.cell(row, col_1 + 1).value = '0.00'
            if not cell_3:
                ws.cell(row, col_1 + 2).value = '0.00'
            if not cell_4:
                ws.cell(row, col_1 + 3).value = '0.00'


# %%
def decryption_sheets(wb, df_matrica, df_avancor, file_fond):
    """Копируем данные из форм с расшифровками и возвращаем список пустых форм"""

    # кол-во строк и столбцов в файле Аванкор
    index_max = df_avancor.shape[0]
    collumn_max = df_avancor.shape[1]

    # Копируем данные расшифровки
    # Выбираем формы: расшифровки разделов
    sheet_2_decryption_names = wb.sheetnames[19:64]
    sheet_2_decryption_names.remove('0420502 Справка о стоимости _56')
    sheet_2_decryption_names.remove('0420502 Справка о стоимости _57')
    # sheet_2_decryption_names.remove('_dropDownSheet')

    form_null = []  # список пустых форм
    for form in sheet_2_decryption_names:
        print(f'{form}')

        ws = wb[form]

        # находим, используя матрицу, раздел в файле Аванкор, соответствующий выбранной форме
        title_1_name = df_matrica.loc[form, 'sheet_1_title']

        # Номер строки с названием раздела в файле Аванкор"""
        title_row = razdel_name_row(df_avancor, title_1_name, index_max, ERRORS, errors_file)

        # находим номер первой строку с данными в файле Аванкор
        data_row = start_data_row(df_avancor, index_max, title_row)

        # находим номер последнке строки с данными в таблице Аванкор"""
        row_end = end_data_row(df_avancor, index_max, data_row)

        if data_row != row_end:
            # список всех номеров строк в таблице Аванкор
            rows_numbers = [x for x in range(data_row, row_end + 1)]
        else:
            # если номера первой и последней строки совпадают, то Раздел пуст и
            # запоминаем название вкладки
            form_null.append(form)
            # переходим к поиску следующего раздела
            continue

        # цифра в последней колонке и номер строки с идектификаторами в таблице XBRL
        # (количество колонок для копирования)
        max_number, id_row = end_col_number(ws)

        # Вставить нужное кол-во строк перед строкой "Итого"
        ws.insert_rows(ws.max_row, amount=(len(rows_numbers) - 1))

        # координаты первой ячейки в таблице XBRL
        col_begin, row_begin = begin_cell(ws, max_number)

        # номера колонок с иденитификаторами в таблице XBRL
        id_cols = id_nombers(col_begin)

        # Номера колонок в таблице Аванкор, за исключением пустых"""
        columns_numbers = find_columns_numbers(df_avancor, collumn_max, max_number, data_row)

        # Копирование данных из таблицы Аванков в таблицу XBRL
        copy_data(ws, df_avancor, rows_numbers, columns_numbers, row_begin, col_begin)

        # Копирование идентификаторов
        # Находим и вставляем идентификаторы в ячейки

        # insert_id(ws, id_row, id_cols, df_id, rows_numbers, row_begin, col_begin, max_number, id_fond)
        insert_id(ws, id_row, id_cols, df_id, rows_numbers, row_begin, col_begin, id_fond)

        # # Сохраняем в файл отчетности xbrl
        # wb.save(file_fond)

    return form_null


# %%

def itogi_sheets(wb_xbrl, df_avancor, df_matrica):
    """ Копируем итоговые данные"""

    # Выбираем вкладки
    sheet_2_itigi = wb_xbrl.sheetnames[6:19]

    for form in sheet_2_itigi:
        print(form)
        # находим, используя матрицу, раздел в файле Аванкор, соответствующий выбранной форме
        title_1_name = df_matrica.loc[form, 'sheet_1_title']

        # координаты первой ячейки с данными в таблице Аванкор
        cell_start = df_matrica.loc[form, 'cell1']
        cell_start_row, cell_start_col = coordinate(cell_start)

        # координаты последней ячейки с данными в таблице Аванкор
        cell_end = df_matrica.loc[form, 'cell_end']
        cell_end_row, cell_end_col = coordinate(cell_end)

        # (количество колонок для копирования) в таблице Аванкор
        max_number = df_matrica.loc[form, 'tbl_col']

        # Номера колонок в таблице Аванкор (за исключением пустых)
        columns_numbers = find_columns_numbers(df_avancor, cell_end_col, max_number, cell_start_row,
                                               data_col=cell_start_col)

        # список всех номеров строк в таблице Аванкор
        rows_numbers = [x for x in range(cell_start_row, cell_end_row + 1)]

        ws_xbrl = wb_xbrl[form]

        # координаты первой ячейки в таблице xbrl
        col_begin, row_begin = begin_cell(ws_xbrl, max_number)

        # Копирование данных из таблицы Аванков в таблицу XBRL
        copy_data(ws_xbrl, df_avancor, rows_numbers, columns_numbers, row_begin, col_begin)


# %%
# ЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖЖ

# %%

# файл с ощибками
errors_file = 'errors.txt'
ERRORS = []

# --------------------------------------------
# загружаем данные из файла c идентификаторами
file_id = r'./Шаблоны/Идентификаторы.xlsx'
df_id = pd.read_excel(file_id, sheet_name=None, header=None)

# %%

# Выбор файла, созданного Аванкор
file_avancor, id_fond = file_open(df_id)
# --------------------------------------------
# %%


# %%

# добавляем к названию файла ошибок идентификатор фонда
errors_file = os.path.splitext(file_avancor)[0] + " - " \
              + os.path.splitext(errors_file)[0] \
              + os.path.splitext(errors_file)[1]
# --------------------------------------------

# %%

# загружвем данные из отчета таблицы Аванкор
df_avancor = load_avancor(file_avancor)

# %%

# загружвем данные из матрицы
df_matrica = load_matrica()

# --------------------------------------------

# кол-во строк и столбцов в файле Аванкор
index_max = df_avancor.shape[0]
collumn_max = df_avancor.shape[1]
# df_avancor.head(3)
# --------------------------------------------
# # Определяем идентификатор фонда их имени файла Аванкор
# id_fond = file_avancor.split('.')[0]
# --------------------------------------------
# название нового файла-отчетности xbrl
print(f'Имя нового файла отчетности....: ', end='')
file_fond_name = asksaveasfilename(title="Имя нового файла отчетности...",
                                   filetypes=(("xlsx files", "*.xlsx"), ("All files", "*.*")))
# Добавляем расширение файла
file_fond_name = file_fond_name + '.xlsx'
# отбрасываем путь к файлу
file_fond = os.path.basename(file_fond_name)
# file_fond = id_fond + '.xlsx'
print(f'{file_fond}')

# --------------------------------------------
# название файла-шаблона
file_shablon = '0420502_0420503_Квартал - 3_1.xlsx'
# file_shablon = '0420502_0420503_Квартал - 3_2.xlsx'
print(f'Используем шаблон: {file_shablon}')
# --------------------------------------------
# Создаем новый файл отчетности xbrl, создав копию шаблона
shutil.copyfile(r'./Шаблоны/' + file_shablon, file_fond_name)

# Загружаем данные из файла таблицы xbrl
wb = openpyxl.load_workbook(filename=file_fond_name)
# wb.sheetnames
# --------------------------------------------
# Записываем во все формы идентификатор фонда
copy_id_fond_to_tbl(wb, id_fond)
# --------------------------------------------

# Копируем данные из форм с расшифровками и возвращаем список пустых форм
form_null = decryption_sheets(wb, df_matrica, df_avancor, file_fond_name)

# Копируем итоговые данные в формы xbrl
itogi_sheets(wb, df_avancor, df_matrica)

# Копируем данные из пояснительных записок!!!!!
sheet_2_zapiski_names = wb.sheetnames[:6]   # список названий листов - пояснительных записок
# Формы:
# 0420502 Пояснительная записка к
# 0420502 Пояснительная записка_2
# 0420502 Пояснительная записка_3
# 0420502 Пояснительная записка_4
# 0420502 Пояснительная записка_5
# 0420502 Пояснительная записка_6

# удаляем лишнюю форму (этой формы нет в Аванкоре)
sheet_2_zapiski_names.remove('0420502 Пояснительная записка_5')

form_zapiski_null = []  # список пустых форм

for form in sheet_2_zapiski_names:
    print(f'{form}')

    ws = wb[form]

    # находим, используя матрицу, раздел в файле Аванкор, соответствующий выбранной форме
    title_1_name = df_matrica.loc[form, 'sheet_1_title']

    # Номер строки с названием раздела в файле Аванкор"""
    title_row = razdel_name_row(df_avancor, title_1_name, index_max, ERRORS, errors_file)

    # находим номер первой строку с данными в файле Аванкор
    data_row = start_data_row(df_avancor, index_max, title_row)

    # находим номер последней строки с данными в таблице Аванкор"""
    row_end = end_data_row(df_avancor, index_max, data_row)

    if df_avancor.loc[row_end - 1, 3] != '2':
        # список всех номеров строк в таблице Аванкор
        rows_numbers = [x for x in range(data_row, row_end)]
    else:
        # Если в соседней ячейки(колонка "C") == "2", то это заголовок таблицы
        # и следовательно форма пустая
        # Запоминаем название вкладки
        form_zapiski_null.append(form)
        # переходим к поиску следующего раздела
        continue

    # цифра в последней колонке и номер строки с идектификаторами в таблице XBRL
    # (количество колонок для копирования)
    max_number = df_matrica.loc[form, 'tbl_col']

    # координаты первой ячейки в таблице XBRL
    cell_start = df_matrica.loc[form, 'cell2']
    cell_start_row, cell_start_col = coordinate(cell_start)

    # Номера колонок в таблице Аванкор, за исключением пустых
    columns_numbers = find_columns_numbers(df_avancor, collumn_max, max_number, data_row, data_col=3)

    # Копирование данных из таблицы Аванков в таблицу XBRL
    copy_data(ws, df_avancor, rows_numbers, columns_numbers, cell_start_row, cell_start_col)
    ERRORS.append(f'"{ws.title}" - вставьте "Идентификатор строки"')

# %%

"""Ручные корректировки"""
# Меняем местами значения ячеек
corrector_scha_ch_1(wb, df_id, id_fond)
# Проставляем нулевые значения в формах
corrector_scha_ch_3to10(wb)
# Вставляем данные о периоде отчетности
corrector_scha_ch_2to12(wb, df_avancor, df_matrica)
# Записываем в формы ФИО подписантов
corrector_scha_56to57(wb, df_matrica, df_avancor, ERRORS, errors_file)
# Вставляем в ячейки с подписантами полное ФИО вместо сокращенного ФИО
corrector_scha_56to57_fio_full(wb, file_id, ERRORS, errors_file)
# id Фонда и реквизиты СпецДепа
corrector_scha_57(wb, df_id, id_fond)
# Копируем кол-во паев с заданной точностью
corrector_scha_13(id_fond, wb, df_avancor, df_id)
# Убираем лишние нули, которые появляются в результате анализа данных
corrector_scha_00(wb)
# Вставляем в форму подробное описание имущества
corrector_scha_51(wb, df_id)

# %%
""" Удаляем пустые формы расшифровки и пояснительных записок"""
for name in (form_null + form_zapiski_null):
    wb.remove(wb[name])

# данные в этой форме в Аванкор отсутствуют - удаляем форму
wb.remove(wb['0420502 Пояснительная записка_5'])

# удаляем формы из Прироста, которые не заполняются
wb.remove(wb['0420503 Отчет о приросте об у_3'])
wb.remove(wb['0420503 Отчет о приросте об у_5'])
wb.remove(wb['0420503 Отчет о приросте об у_6'])

# %%

# Сохраняем результаты в файл отчетности xbrl
try:
    # os.chdir (os.path.dirname (file_fond_name))
    wb.save(file_fond_name)
    print('-------------------ГОТОВО!!!----------------------')

    # messagebox.showinfo("Конвертер", "Файл создан!")

except PermissionError:
    ERRORS.append(f'{"=" * 100}\n'
                  f'!!!ФАЙЛ НЕ СОЗДАН!!! \n'
                  f'ОШИБКА ДОСТУПА К ФАЙЛУ '
                  f'(файл открыт в другой программе - закройте файл!)\n'
                  f'{"=" * 100} ')

# %%

# Записываем ошибки
write_errors(ERRORS, errors_file)

# %%
# Открываем файл с ошибками в Блокноте
# Блокнот
notepad = r'%windir%\system32\notepad.exe'
file = notepad + ' ' + errors_file
os.system(file)

# %%
