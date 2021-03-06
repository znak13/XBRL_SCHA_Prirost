# import sys
# from module.functions import coordinate
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import column_index_from_string  # 'B' -> 2
from openpyxl.styles import colors
from openpyxl.styles import Font

from module.globals import error_txt
global log


# %%
def find_parametr(ws, row_begin, col):
    """ Находим название столбца при определении возможной ошибки """

    for row in range(1, 10):
        cell = ws.cell(row_begin - row, col).value

        # В ячейке цифра(номер столбца) или
        # в пояснительной записке №3 вместо номера столбца укзано "Содержание"
        if str(cell).isdigit() or \
                (str(cell) == "Содержание" and str(ws.cell(3, 1).value).endswith('SR_0420502_PZ_inaya_inf')):
            # название столбца находится на строчку выше
            row_param = row_begin - row - 1
            cell = ws.cell(row_param, col).value
            return cell

        # в пояснительной записке №2 нет номера столбца
        # в пояснительной записке №5 нет номера столбца: 'Сумма'
        # в "0420502 Справка о стоимости _56" нет номера столбца: 'Содержание'
        if str(cell).startswith('Сведения о событиях') or \
                str(cell) == 'Сумма' or \
                (str(cell) == 'Содержание' and str(ws.cell(3, 1).value).endswith('SR_0420502_Podpisant')):
            # название столбца находится в этой ячейке
            row_param = row_begin - row
            cell = ws.cell(row_param, col).value
            return cell


# %%
def check_errors(ws, cell_avancor, row_begin, r, c, col_begin):
    """ Проверка на наличие ошибок """

    # Наименование колонки в таблице xbrl
    col_name = find_parametr(ws, row_begin, col_begin + c)

    if (str(cell_avancor).startswith("Не установлен")
        or str(cell_avancor) == 'nan'
        or str(cell_avancor) == '-') \
            and col_name != 'Примечание':
        log.error(f'"{ws.title}"; '
                  f'строка({row_begin + r}), '
                  f'колонка({c + 1}) --> '
                  f'параметр: "{col_name}"'
                  f' ==> "{cell_avancor}"')


# %%
def empty_cell(ws, cellBegin, cellEnd):
    """ Проверяем является ли ячейка пустой"""

    rowBegin, colBegin = coordinate_to_tuple(cellBegin)
    rowEnd, colEnd = coordinate_to_tuple(cellEnd)

    drow = rowBegin + (rowEnd - rowBegin) + 1
    dcol = colBegin + (colEnd - colBegin) + 1

    for row in range(rowBegin, rowBegin + (rowEnd - rowBegin) + 1):
        for col in range(colBegin, colBegin + (colEnd - colBegin) + 1):
            cellData = ws.cell(row, col).value
            if not cellData or \
                    str(cellData).startswith('Не установлен') or \
                    cellData == 'None' or \
                    cellData != cellData or \
                    cellData == error_txt:
                red_error(ws.cell(row, col))
                log.error(f'"{ws.title}" --> пустая ячейка "{get_column_letter(col) + str(row)}"')


# %%
def red_error(cell):
    """Окрашивание ошибки в красный цвет"""
    # from openpyxl.styles import colors
    # from openpyxl.styles import Font

    cell.value = error_txt
    # красный цвет
    color_font = colors.Color(rgb='FFFF0000')
    cell.font = Font(color=color_font)


# %%
def id_errors(ws, columns: (list or tuple), row_begin: int = 11):
    """Проверяем ячейки с идентификаторами на предмет отсутствия идентификатора.
    Если ячейка пустая, то это ошибка, т.к. обязательно должен быть указан идентификатор."""

    for row in range(row_begin, ws.max_row):
        for col in columns:
            cell = ws.cell(row, column_index_from_string(col))
            if cell.value in [None, 'None', error_txt]:
                red_error(cell)
                log.error(f'"{ws.title}", строка({row}), колонка({col}) --> отсутствует Идентификатор')



# %%
def txt_compare(txt1: str, txt2: str) -> bool:
    """Сравнение друх строк текста, исключая пробелы"""

    if txt1.replace(' ', '') == txt2.replace(' ', ''):
        return True
    return False


# ==================================================================================

if __name__ == "__main__":
    pass

    txt1 = ' q w e r t y'
    txt2 = 'qw erty'
    print(txt_compare(txt1, txt2))
