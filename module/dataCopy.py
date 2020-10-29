import openpyxl
from openpyxl.styles import Font

from module.dataCheck import check_errors
from module.dataCheck import red_error
from module.analiz_data import analiz_data_all
from module.functions import coordinate, find_columns_numbers, \
    razdel_name_row, start_data_row, end_data_row

from module.globals import *
global log

def copy_data(ws, df_avancor, rows_numbers, columns_numbers, row_begin, col_begin):
    """ Копирование данных из таблицы Аванкор в таблицу XBRL """

    for r, row in enumerate(rows_numbers):
        for c, col in enumerate(columns_numbers):
            cell_avancor = df_avancor.loc[row, col]
            check_errors(ws, cell_avancor, row_begin, r, c, col_begin)
            if cell_avancor != '-' and \
                    cell_avancor != "x" and \
                    cell_avancor != 0 and \
                    str(cell_avancor) != 'nan':
                ws.cell(row_begin + r, col_begin + c).value = \
                    analiz_data_all(df_avancor.loc[row, col])


# %%
def copy_id_fond_to_tbl(ws, id_fond):
    """Записываем в форму идентификатор фонда"""

    # идентификатор фонда содержится во всех формах, кроме:
    # 0420502 Справка о стоимости чис - SR_0420502_R1
    # 0420502 Справка о стоимости _56 - SR_0420502_Podpisant
    # 0420502 Справка о стоимости _57 - SR_0420502_Podpisant_spec_dep

    # Во всех формах идентификатор фонда содержится
    # в строке "5" и в крайней правой колонке,
    # а текст в ячейке начинается с 'Z= Идентификатор АИФ ПИФ-'

    cell_id = ws.cell(row=5, column=ws.max_column)
    # info = 'Z= Идентификатор АИФ ПИФ-'
    if fondIDtxt in cell_id.value:
        cell_id.value = fondIDtxt + id_fond


# %%
def id_serch(ws, sheet_name_id, row_i, col_begin, id_fond, df_id):
    """ Определение идентификатора
        :ws: форма в отчетности xbrl
        :sheet_name_id: название листа с в фойле с идентификаторами
        :row_i: текущая строка в в таблице xbrl
        :col_begin: начальная колонка в таблице xbrl
        :id_fond: идентификатор фонда
        :return 'id_is': искомый идентификатор
    """

    def round_id(df_id_ws, n, id_fond=None):

        for col in range(col_begin, ws.max_column + 1):
            id_priznak = ws.cell(row_i, col).value if id_fond is None else id_fond

            # Убираем пробелы в идентификаторах
            # (так значния будут сравниваться корректнее)
            if id_priznak:
                id_priznak = id_priznak.replace(' ', '')

            # Убираем пробелы в столбце файла с идентификаторами (в каждом элементе)

            # и запоминаем новые значения (уже без пробелов)
            # df_id_ws[n] = df_id_ws[n].apply(lambda x: x.replace(' ', '')) # (1)вариант без новой переменной

            # (заводим новую переменную, т.к. если оставить (1)вариант,
            # то появляется предупреждение о возможной некорректрой работе)
            id_names = df_id_ws[n] # значения n-го столбца в файле с идентификаторами
            id_names = id_names.apply(lambda x: x.replace(' ', ''))

            # Список признаков
            # df_priznak = df_id_ws[df_id_ws[n] == id_priznak]              # (1)вариант без новой переменной
            df_priznak = df_id_ws[id_names == id_priznak]

            # Если находим признак в строке, то прерываем перебор колонок
            # if id_priznak in df_priznak[n].to_list():                     # (1)вариант без новой переменной
            # (перед сравнением удаляем лишние пробелы)

            if id_priznak in [x.replace(' ', '') for x in df_priznak[n].to_list()]:
                break


        return df_priznak, id_priznak

    df_priznak = df_id[sheet_name_id]   # текущий df
    # Меняем тип всех данных на "str"
    # (это необходимо для корректного сравнения в дальнейшем)
    df_priznak = df_priznak.astype(str)

    # df_priznak = df_id_ws  # текущий df
    df_priznak_list = []  # список найденных df
    id_is = None  # искомый идентификатор

    n = 1
    while n < df_priznak.columns.size:
        df_priznak, id_priznak = round_id(df_priznak, n)
        # допоминаем список c 'df_priznak'
        df_priznak_list.append(df_priznak)
        if df_priznak.index.size == 1:  # одна строчка
            # идентификатор - это превый элемент в 'df_priznak'
            id_is = df_priznak[0].values[0]
            break
        elif df_priznak.index.size == 0:  # если список признаков пуст, то ищем по 'id_fond'
            if sheet_name_id == 'Биржа':  # идентификатор не найден
                id_is = df_priznak.loc[1, 0]
                # id_is = no_birzha
            elif sheet_name_id == 'Банковский счет':
                # берем предыдущий 'df_priznak' () из 'df_priznak_list'
                # (отнимаем "2",т.к. в списке 'df_priznak_list' элементы начинаются с "0")
                df_priznak, id_priznak = round_id(df_priznak_list[n - 2], n + 1, id_fond=id_fond)
                # идентификатор - это превый элемент в 'df_priznak'
                try:
                    id_is = df_priznak[0].values[0]
                except Exception:
                    log.error(f'в файле "Идентификаторы" не указан расчетный счет фонда')
            else:
                id_is = 'ошибка'
                log.error(f'"{ws.title}" --> не определен "{df_priznak.loc[0, 0]}", строка:{row_i} ')

            break
        else:
            n += 1

    if id_is:
        return id_is
    else:
        id_is = 'ошибка'
        log.error(f'Идентификатор не найден: '
                  f'"{ws.title}", "{df_priznak.loc[0, 0]}", строка:{row_i}')

        return id_is


# %%

def insert_id(ws, id_row, id_cols, df_id, rows_numbers, row_begin, col_begin, id_fond):
    """ Находим и вставляем идентификаторы в ячейки"""
    # id_cols - номера колонок с иденитификаторами в таблице XBRL

    # словарь ВСЕХ названий идетификаторов (полные) и имен вкладок (сокращенные)
    list_id = {(df_id[k].loc[0, 0]): k for k in df_id.keys()}
    # если ошибка в этой строке, то проверить названеи вкладок в файле с идентификаторами
    # ...

    # список идентификаторов на листе
    id_on_list = [ws.cell(id_row, col).value for col in id_cols]
    # находим признаки идентификатора в строке
    for i, id_name in enumerate(id_on_list):  # полные названия идентификаторов
        # название листа в файле с идентификатрами (сокращенные названия идентификаторов)
        sheet_name_id = list_id[id_name]

        for row in range(len(rows_numbers) - 1):
            row_i = row_begin + row
            id_is = id_serch(ws, sheet_name_id, row_i, col_begin, id_fond, df_id)

            # записываем название идентификатора в таблицу xbrl
            cell = ws.cell(row_i, id_cols[i])
            cell.value = id_is

            # отмечаем красным цветом ошибку
            if id_is == "ошибка":
                red_error(cell)
                # # красный цвет
                # color_font = openpyxl.styles.colors.Color(rgb='FFFF0000')
                # ws.cell(row_i, id_cols[i]).font = Font(color=color_font)


# %%

if __name__ == "__main__":
    # красный цвет
    # color_font = openpyxl.styles.colors.Color(rgb='FFFF0000')
    # ws_xbrl[cell.coordinate].font = Font(color=color_font)
    pass


    # ----------------------------------------------------------
    from module.data_load import load_info_from_files

    # Папка с шаблонами
    dir_shablon = r'../Шаблоны/'
    path_to_report = '.' + dir_reports + '/2020_09/'
    file_Avancore_scha = 'ЗПИФ_Азов_сча.xlsx'
    file_new_name = '1122334455.xlsx'

    df_identifier, df_avancor, wb = load_info_from_files(dir_shablon,
                                                         fileID,
                                                         path_to_report,
                                                         file_Avancore_scha,
                                                         file_new_name)

    # ----------------------------------------------------------

    # id_is = id_serch(ws, sheet_name_id, row_i, col_begin, id_fond, df_id)