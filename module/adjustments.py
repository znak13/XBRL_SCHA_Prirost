""" Корректировки содержания форм отчетности после копирования данных из файла Аванкор"""

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string  # 'B' -> 2
from module.functions import coordinate
from module.functions import codesSheets, sheetNameFromUrl
from module.functions import razdel_name_row, start_data_row, end_data_row, find_columns_numbers
from module.functions import pai_search
from module.analiz_data import analiz_data_data, toFixed
from module.dataCheck import red_error
from module.globals import *

# from module.data_load import load_pif_info

global log


def corrector_scha_01(wb, id_fond, shortURL=None):
    """ Меняем местами значения ячеек и добавляем id фонда и УК"""
    # Форма:
    # 0420502 Справка о стоимости чис   SR_0420502_R1

    sheetName = sheetNameFromUrl(codesSheets(wb), shortURL)
    ws_change = wb[sheetName]

    # Меняем значения ячеек, т.к. в таблице xbrl ячейки расположены
    # в ином порядке, чем в файле, созданном Аванкор
    ws_change.cell(10, 5).value, ws_change.cell(10, 6).value = \
        ws_change.cell(10, 6).value, ws_change.cell(10, 5).value
    ws_change.cell(10, 5).value, ws_change.cell(10, 7).value = \
        ws_change.cell(10, 7).value, ws_change.cell(10, 5).value

    # Записываем в форму идентификаторы: фонда и УК
    ws_change.cell(10, 1).value = id_fond
    # ws_change.cell(10, 2).value = df_id['УК АИФ ПИФ'].loc[1, 0]
    ws_change.cell(10, 2).value = uk_ogrn


# %%
def corrector_scha_03to10(wb, shortURL=None):
    """ Проставляем нулевые значения в формах: ч_3 - ч_10 """

    # Формы:
    # 0420502 Справка о стоимости ч_3    SR_0420502_R3_P1
    # 0420502 Справка о стоимости ч_4    SR_0420502_R3_P4
    # 0420502 Справка о стоимости ч_5    SR_0420502_R3_P2
    # 0420502 Справка о стоимости ч_6    SR_0420502_R3_P3
    # 0420502 Справка о стоимости ч_7    SR_0420502_R3_P5
    # 0420502 Справка о стоимости ч_8    SR_0420502_R3_P6
    # 0420502 Справка о стоимости ч_9    SR_0420502_R3_P7
    # 0420502 Справка о стоимости _10    SR_0420502_R3_P8

    def insert_00(ws):
        """ Проставляем нулевые значения в форме """

        row_begin = 10
        col_1 = 3  # номер первой колонки с данными

        for row in range(row_begin, ws.max_row + 1):
            cell_1 = ws.cell(row, col_1).value
            cell_2 = ws.cell(row, col_1 + 1).value
            cell_3 = ws.cell(row, col_1 + 2).value
            cell_4 = ws.cell(row, col_1 + 3).value

            # Форматируем ячейки
            for i in range(4):
                ws.cell(row, col_1 + i).alignment = Alignment(horizontal='right')

            if cell_1 or cell_2:  # Если есть данные в одной из первых ячеек
                if not cell_1:  # если нет в первой
                    ws.cell(row, col_1).value = '0.00'
                if not cell_2:  # если нет во второй
                    ws.cell(row, col_1 + 1).value = '0.00'
                if not cell_3:  # если нет в третьей
                    ws.cell(row, col_1 + 2).value = '0.00'
                if not cell_4:  # если нет в червертой
                    ws.cell(row, col_1 + 3).value = '0.00'

    sheet = sheetNameFromUrl(codesSheets(wb), shortURL)
    ws = wb[sheet]
    insert_00(ws)


# %%
def corrector_scha_03to12_(wb_xbrl, df_avancor, cell_period, shortURL=None):
    """ Вставляем данные о периоде отчетности """
    # Формы:
    # 0420502 Справка о стоимости ч_3   SR_0420502_R3_P1
    # 0420502 Справка о стоимости ч_4   SR_0420502_R3_P4
    # 0420502 Справка о стоимости ч_5   SR_0420502_R3_P2
    # 0420502 Справка о стоимости ч_6   SR_0420502_R3_P3
    # 0420502 Справка о стоимости ч_7   SR_0420502_R3_P5
    # 0420502 Справка о стоимости ч_8   SR_0420502_R3_P6
    # 0420502 Справка о стоимости ч_9   SR_0420502_R3_P7
    # 0420502 Справка о стоимости _10   SR_0420502_R3_P8
    # 0420502 Справка о стоимости _11   SR_0420502_R3_P9
    # 0420502 Справка о стоимости _12   SR_0420502_R4

    period_begin = analiz_data_data(df_avancor.loc[18, 1])
    period_end = analiz_data_data(df_avancor.loc[18, 5])
    sheet_xbrl_period = sheetNameFromUrl(codesSheets(wb_xbrl), shortURL)

    ws_period = wb_xbrl[sheet_xbrl_period]
    # cell_period = df_matrica.loc[shortURL, 'cell_period']
    ws_period[cell_period].value = period_begin + ', ' + period_end


# %%
def corrector_scha_13_(id_fond, ws, df_avancor):
    """ Копируем количество паев с нужной точностью """
    # форма:
    # 0420502 Справка о стоимости _13

    # Точность указания кол-ва паев
    fix = pai_search(id_fond)
    # или так:
    # wb_pif_info = load_pif_info(file_name=pif_info,
    #                                 path_2file=dir_shablon)
    # sheet_name = id_fond
    # ws_pif_info = wb_pif_info[sheet_name]
    # col_fIx = 'B'
    # row_fix = 2
    # cell = col_fIx + str(row_fix)
    # fix = ws_pif_info[cell].value

    today_row, today_col = coordinate('I144')
    previous_row, previous_col = coordinate('K144')

    today = toFixed(df_avancor.loc[today_row, today_col], digits=fix)
    previous = toFixed(df_avancor.loc[previous_row, previous_col], digits=fix)

    ws.cell(10, 3).value = today
    ws.cell(10, 4).value = previous

    # Форматируем ячейки
    for row in range(9, ws.max_row + 1):
        for col in range(3, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(horizontal='right')


# %%

def corrector_Podpisant_3_v2(ws_xbrl, id_fond):
    # Загрузка файла
    ws_SD = pd.read_excel(dir_shablon + pif_info,
                          sheet_name=id_fond,
                          index_col=0,
                          header=None)
    sd_name = ws_SD.loc['СД', 1]
    sd_inn = ws_SD.loc['СД_ИНН', 1]
    sd_ogrn = ws_SD.loc['СД_ОГРН', 1]
    sd = [sd_name, sd_inn, sd_ogrn]

    for n, col in enumerate(range(3, 6)):
        ws_xbrl.cell(8, col).value = sd[n]


# %%
def depozitID(df_avancor, avancoreTitle, max_number=10):
    """ При наличии депозитов, в файле-Аванкор нужно обязательно указывать номер счета депозита"""
    # Раздел - '1.2. Денежные средства на счетах по депозиту в кредитных организациях'
    # max_number - количество колонок в таблице

    # row_data, row_end - номер первой и последней строки с данными в таблице Аванкор
    row_data, row_end = data_search(df_avancor, avancoreTitle)

    # кол-во столбцов в файле Аванкор
    collumn_max = df_avancor.shape[1]
    # Если номера первой и последней строки НЕ совпадают, то Раздел содержит данные
    if row_data != row_end:
        # Номера колонок в таблице Аванкор, за исключением пустых"""
        columns_numbers = find_columns_numbers(df_avancor, collumn_max, max_number, row_data)
        # номер последней колонки в таблице ('Примечание')
        col_prim = columns_numbers[-1]
        for row in range(row_data, row_end):
            if df_avancor.loc[row, col_prim] == '-':
                log.error(f'файл: Аванкор(СЧА) - раздел:"{avancoreTitle}", строка:({row}) '
                          f'- в колонке "Примечание" укажите номер счета депозита')


# %%
def data_search(df_avancor, avancoreTitle):
    """ Поиск в таблице-Аванкор наличия строк с данными"""

    # Кол-во строк и столбцов в файле Аванкор
    index_max = df_avancor.shape[0]
    collumn_max = df_avancor.shape[1]

    # Номер строки с названием раздела в файле Аванкор"""
    title_row = razdel_name_row(df_avancor, avancoreTitle, index_max)

    # находим номер первой строку с данными в файле Аванкор
    row_data = start_data_row(df_avancor, index_max, title_row)

    # находим номер последнке строки с данными в таблице Аванкор"""
    row_end = end_data_row(df_avancor, index_max, row_data)

    return row_data, row_end


def corrector_00(ws, row: int = None, col: str = None):
    """ Убираем лишние '.00' в конце строки,
    которые могут появиться после копирования данных"""
    # (например при копировании ИНН: "7705373131" => "7705373131.00" )
    # row: int - начальная строка столбца с данными
    # col: str - колонка столбца с данными

    col = column_index_from_string(col)
    for r in range(row, ws.max_row + 1):
        cell = ws.cell(r, col)
        value = cell.value
        # в тексте есть '.00'
        if value.endswith('.00'):
            value = value[:-3]
            cell.value = value


def corrector_00_v2(ws, *colls, row: int = 11):
    """ Убираем лишние '.00' в конце строки,
    которые могут появиться после копирования данных"""
    # (например при копировании ИНН: "7705373131" => "7705373131.00" )
    # row: int - начальная строка столбца с данными
    # collls: str - колонка или колонки с данными

    for col in colls:
        for r in range(row, ws.max_row + 1):
            cell = ws.cell(r, column_index_from_string(col))
            value = cell.value
            if value:  # если ячейка не пустая
                if value.endswith('.00'):  # в конце текста есть '.00'
                    value = value[:-3]
                    cell.value = value


def copy_cells_one2one(ws, row_begin, col_to, col_from,
                       del_old_sell=False, id_fond=''):
    if id_fond:
        id_fond = '_(' + id_fond + ')'

    for row in range(row_begin, ws.max_row):

        ws.cell(row, column_index_from_string(col_to)).value = \
            str(ws.cell(row, column_index_from_string(col_from)).value) + id_fond

        if del_old_sell:  # стираем значение в ячейке, из которой копировали
            ws.cell(row, column_index_from_string(col_from)).value = ""


def copy_bank_account(id_fond, wb_pif_info, ws, row_begin, col_to, col_from):
    """Копирование банковского счета"""

    # from openpyxl.styles import colors
    # from openpyxl.styles import Font

    sheet_name = id_fond
    # title = 'тип счета'
    ws_pif_info = wb_pif_info[sheet_name]
    col_pif = 'B'

    for row_ws in range(row_begin, ws.max_row):
        prim = str(ws.cell(row_ws, column_index_from_string(col_from)).value)

        # если в ячейке цифра со знаками после запятой, то берем только целую часть
        # (может появиться после анализа данных при копировании из Аванкор )
        prim = prim.split('.')[0]
        # print(prim)

        cell_to = ws.cell(row_ws, column_index_from_string(col_to))
        if prim and prim == prim and prim != 'None':
            for row_pif in range(1, ws_pif_info.max_row):
                # номер счета
                bank_account = str(ws_pif_info.cell(row_pif, column_index_from_string(col_pif)).value)
                # если номер счета заканчивается на 'prim'
                if bank_account.endswith(prim):
                    # копируем номер счета из файла "pif_info"
                    cell_to.value = bank_account

                    # стираем значение в ячейке "Примечание"
                    ws.cell(row_ws, column_index_from_string(col_from)).value = ""

        else:
            red_error(cell_to)


def make_id(txt: str, start=1, end=2) -> str:
    """Создание идентификатора из текста"""

    start = start - 1
    id_txt = txt.split()[start:end]
    return '_'.join(id_txt)


def copy_path_of_cells(ws, row_begin, col_to, col_from):
    for n, row in enumerate(range(row_begin, ws.max_row), 1):
        txt_from = ws.cell(row, column_index_from_string(col_from)).value
        id_txt = make_id(txt_from) + '_' + str(n)
        ws.cell(row, column_index_from_string(col_to)).value = id_txt


def copy_hash_of_cells(id_fond, ws, row_begin, col_to, col_from,
                       word_start=0, word_end=2,
                       delta=False, fio=False, dogovor_n=False, fond_name=False):
    """Строим идентификатор из строки"""
    # delta - добавляем индекс после идентификатора
    # (для исключения появления одинаковых идентификаторов)

    for n, row in enumerate(range(row_begin, ws.max_row), 1):
        txt_from = str(ws.cell(row, column_index_from_string(col_from)).value)
        txt_from_list = txt_from.split()

        # если слов в строке меньше, чем 'word_end', то уменьшаем 'word_end'
        if len(txt_from_list) < word_end:
            word_end = len(txt_from_list)

        # если последнее слово состоит менее чем из 3-х букв, то исключаем его
        if len(txt_from_list[word_end - 1]) < 3:
            word_end -= 1

        txt_list = txt_from_list[word_start:word_end]

        # # удаляем точку в конце последнего слова
        # # (актуально при копировании паспортных данных)
        # if len(txt_list) == 2:
        #     if txt_list[1].endswith('.'):
        #         txt_list[1] = txt_list[1][:-1]

        # # если id состоит из ФИО, то находим серию и номер паспорта
        # if fio:
        #     txt_list = pasport_from_str(' '.join(txt_list))

        first_word = '_'.join(txt_list)
        id_txt = ""

        # ...если нужно, то добавляем порядковый номер требования
        if delta:
            id_txt = '(' + str(n) + ')'

        # Если Основания начинается с "Налоговый кодекс"
        nk = 'Налоговый кодекс'
        if txt_from.startswith(nk):
            # добавляем "nk"
            id_txt = id_txt + '_' + nk
            # и добавляем только 'id_fond'

        else:  # продолжаем анализ

            # # ...если нужно вставить Фамилию
            # if fio:
            #     initials = fio_initials(ws.cell(row, column_index_from_string(fio)).value)
            #     if id_txt:
            #         id_txt = id_txt + '_' + initials
            #     else:
            #         id_txt = initials

            # добавляем 'first_word'
            id_txt = id_txt + '_' + first_word

            # ...если нужно, то ищим номер договора:
            if dogovor_n:
                num = '№'
                # если последнее слово - это '№', то исключаем его,
                # т.к. '№' будет вставлен на следующем шаге
                if txt_from_list[word_end - 1].startswith(num):
                    word_end -= 1

                nomber = find_nomber(txt_from, n=num)
                if nomber:  # если найден номер, то

                    if (num + first_word) != str(nomber):
                        # строка состоит их нескольких слов
                        id_txt = id_txt + '_' + str(nomber)
                    else:
                        # строка состоит их одного слова, это - 'nomber'.
                        # в этом случае, для исключения задвоения,
                        # не прописываем 'first_word'
                        # и добавляем слово 'Договор №'
                        id_txt = '(' + str(n) + ')_' + \
                                 'Договор_' + str(nomber)

        # ...если нужно, добавляем 'id_fond'
        if fond_name:
            id_txt = id_txt + '_' + '(' + id_fond + ')'

        ws.cell(row, column_index_from_string(col_to)).value = id_txt
        # + '_(' + str(hash(txt_from)) + ')'


def find_nomber(txt, n='№'):
    """Поиск в строке фрагмента '№'(номер договора)"""
    # № - элемент, с которого начинается искомый fragment
    stop = False
    fragment = None
    txt_list = txt.split()
    for i, word in enumerate(txt_list):
        if word.startswith(n):
            # состоит только из '№'
            if len(txt_list[i]) == 1:
                fragment = n + txt_list[i + 1]
                stop = True
                break
            else:
                fragment = txt_list[i]
                stop = True
                break
    # если во всех словах не найден "№"
    if not stop:
        for i, word in enumerate(txt_list):
            # слово состоит не только из букв
            # или не заканчивается на точку
            # или не является датой
            if not word.isalpha() and \
                    not word.endswith('.') and \
                    len(word.split('.')) < 3:
                fragment = n + txt_list[i]
                break

    return fragment


def fio_initials(txt):
    """ Фамилие и инициалы физ.лица"""
    # используем только первые три слова (ФИО или ФИ)
    txt_list = txt.split()[:3]
    inicials = ''
    if len(txt_list) < 2:
        log.error(f'ФИО физ.лица: "{txt}" состоит только из одного слова.')
    for i in range(1, len(txt_list)):
        inicials += txt_list[i][0].upper() + '.'
    return txt_list[0] + ' ' + inicials


def copy_birzha_id(ws, row_begin, col_to, col_from):
    """Копирование идентификатора Биржи"""

    for n, row in enumerate(range(row_begin, ws.max_row), 1):
        txt_from = ws.cell(row, column_index_from_string(col_from)).value
        if not txt_from:
            ws.cell(row, column_index_from_string(col_to)).value = no_birzha
        else:
            ws.cell(row, column_index_from_string(col_to)).value = \
                ws.cell(row, column_index_from_string(col_from)).value


def corrector_depozit_type(ws, col='I', row=11):
    """ Добавляем пробел ' ' после 'Да' или 'Нет' """

    # 0420502 Справка о стоимости чистых активов,
    # в том числе стоимости активов (имущества),
    # акционерного инвестиционного фонда (паевого инвестиционного фонда)
    # Расшифровки раздела 3 «Активы». Подраздел 1.2.
    # Денежные средства на счетах по депозиту в кредитных организациях

    # 0420502 Справка о стоимости _15

    for r in range(row, ws.max_row):
        cell = ws.cell(r, column_index_from_string(col))
        if not cell.value.endswith(' '):  # в конце текста нет пробела
            cell.value = cell.value + ' '


# %%
# def pasport_from_str(pasport: str):
#     """Найти серию и номер паспорта в тексте"""
#
#     txt_lst = pasport.split()
#     seriya = ''
#     nomer = ''
#     for word in txt_lst:
#
#         # удаляем точку в конце строки
#         word = word.rstrip('.')
#
#         # слово состоит из цифр и
#         # серия паспорта не определена либо состоит только из 2-х знаков
#         if word.isdigit() and (not seriya or len(seriya) == 2):
#             seriya = seriya + word
#             continue
#         # слово состоит из цифр и серия паспорта определена
#         if word.isdigit() and seriya:
#             nomer = word
#             break
#
#     # return '_' + pas_1 + '_' * bool(pas_2) + pas_2
#     return [seriya, nomer]


def pasport_from_str(pasport: str):
    """Найти серию и номер паспорта в тексте"""

    txt_lst = pasport.split()
    seriya = ''
    nomer = ''
    for word in txt_lst:
        # удаляем точку и запятую в конце строки
        word = word.rstrip('.')
        word = word.rstrip(',')
        # слово состоит из цифр и
        # серия паспорта не определена либо состоит только из 2-х знаков
        if word.isdigit() and (not seriya or len(seriya) == 2):
            seriya = seriya + word
            continue
        # слово состоит из цифр и серия паспорта определена
        if word.isdigit() and seriya:
            nomer = word
            break
    return seriya + '_' * bool(nomer) + nomer


def id_fiz_face(ws, row_begin, col_fio, col_pas, col_id):
    """Формирование id физ.лица """
    for row in range(row_begin, ws.max_row):
        fio = str(ws.cell(row, column_index_from_string(col_fio)).value)
        pas = str(ws.cell(row, column_index_from_string(col_pas)).value)
        id_fl = fio_initials(fio) + '_' + pasport_from_str(pas)
        ws.cell(row, column_index_from_string(col_id)).value = id_fl


def find_nomber_2(txt):
    """Поиск в строке фрагмента '№'(номер договора)"""
    fragment = ''
    txt_list = txt.split()
    for i, word in enumerate(txt_list):
        # слово начинается с "№" ("№12/34" или "№")
        if word.startswith('№'):
            # состоит только из одного символа ('№')
            if len(word) == 1:
                # не является последним словом в строке
                if i != len(txt_list) - 1:
                    # если после '№' нет слова 'от'
                    if txt_list[i + 1] != 'от':
                        # запоминаем слово, следующее за "№"
                        fragment = txt_list[i + 1]
                        break
                # является последним словом в строке и состоит только из одного символа "№"
                else:
                    break
            # между "№" и номером нет пробела ('№12/34')
            else:
                fragment = word[1:]
                break
    # если во всех словах не найден "№", то догово без номера
    if not fragment:
        fragment = "б/н"

    return '№' + fragment


def find_dokumentName(txt):
    """Поиск названия документа"""
    txt_list = txt.split()
    if len(txt_list) < 2:
        log.error(f'Внимание: название документа "{txt}" содержит всего одно слово.')
    if txt_list[0].lower() == 'налоговый':
        # если строка начинается с "Налоговый", то речь идет о "Налоговый кодекс"
        # в этом случае возвращаем два первых слова
        return txt_list[0].title() + '_' + txt_list[1].lower()
    else:
        # Возвращаем первое слово в строке
        return txt_list[0].title()


def id_osnovaniya(txt):
    """Идентификатор основания возникновения ... задолженности"""
    name = find_dokumentName(txt)
    nomber = find_nomber_2(txt)
    if name.lower() == 'налоговый_кодекс':
        return name
    return name + '_' + nomber


def id_osnovaniya_zadilzhennosti(id_fond, ws, row_begin, col_to, col_from):
    """Формирование id основания возникновения ... задолженности"""

    for n, row in enumerate(range(row_begin, ws.max_row), 1):
        txt = str(ws.cell(row, column_index_from_string(col_from)).value)
        # ws.cell(row, column_index_from_string(col_to)).value = \
        #     id_osnovaniya(txt) + '_' + '(' + id_fond + ')' + '(' + str(n) + ')'
        _id_ = id_osnovaniya(txt)

        # если текст в колонке встречается более одного раза,
        # то к идентификатору добавляем индекс
        index = find_copy(ws, row_begin, col_to, _id_)

        ws.cell(row, column_index_from_string(col_to)).value = \
            _id_ + '_' * bool(index) + index
            # _id_ + '_' + '(' + id_fond + ')' + index


def find_copy(ws, row_begin, col, txt):
    """Поиск ячеек в столбце:'col', которые начинаются с 'txt'.
    Возвращаем строку с количеством повторений либо пустую строку, если повторений нет"""
    count = 1
    for row in range(row_begin, ws.max_row):
        cell = str(ws.cell(row, column_index_from_string(col)).value)
        if cell.startswith(txt):
            count += 1
    if count > 1:
        return '(' + str(count) + ')'
    return ''

def id_den_treb(id_fond, ws, row_begin, col_to, col_from):
    """ Формирование id денежного требования"""
    for row in range(row_begin, ws.max_row):
        dogovr_number = str(ws.cell(row, column_index_from_string(col_from)).value)
        _id_ = 'Договор' + '_' + dogovr_number

        # если текст в колонке встречается более одного раза,
        # то к идентификатору добавляем индекс
        index = find_copy(ws, row_begin, col_to, _id_)

        ws.cell(row, column_index_from_string(col_to)).value = \
            _id_ + '_'*bool(index) + index
            # _id_ + '_' + '(' + id_fond + ')' + index

def id_dogovor(id_fond, ws, row_begin, col_to, col_from):
    """Формирование id договора"""
    id_den_treb(id_fond, ws, row_begin, col_to, col_from)



# %%

if __name__ == "__main__":
    pass

    # t1 = 'Паспорт иностранного гражданина, 07732966. Выдан Министерством внутренних дел, 16.11.2011 г. '
    # t2 = 'Паспорт гражданина РФ, 6012 075791. Выдан: Отделом УФМС России по Ростовской области в городе Новочеркасске, 07.11.2011'
    # t3 = 'Паспорт гражданина РФ, 12  34 075791. Выдан: Отделом УФМС России по Ростовской области в городе Новочеркасске, 07.11.2011'
    #
    # print (id_from_pasport(t1))

    # id_fond = 'eeeee'
    # txt = 'договор №345345 от'
    # print(id_osnovaniya(txt))
