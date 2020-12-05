""" Формирование форм Прирост"""

import module.functions as fun
import module.dataCopy as dcop
import module.dataCheck as dCheck

from openpyxl.styles import Alignment
from module.analiz_data import analiz_data_all

global log
# Формы Прирост:
# 0420503 Отчет о приросте об уме  -1-   SR_0420503_R1
# 0420503 Отчет о приросте об у_2  -2-   SR_0420503_R2
# 0420503 Отчет о приросте об у_3  -3-   SR_0420503_R3_5   - не заполняется
# 0420503 Отчет о приросте об у_4  -4-   SR_0420503_R3
# 0420503 Отчет о приросте об у_5  -5-   SR_0420503_R4_5   - не заполняется
# 0420503 Отчет о приросте об у_6  -6-   SR_0420503_R4     - не заполняется
# 0420503 Отчет о приросте об у_7  -7-   SR_0420503_podpisant
# 0420503 Отчет о приросте об у_8  -8-   SR_0420503_podpisant_spec_dep


def prirost(wb):
    # Удаляем формы из Прироста, которые не заполняются

    # словарь: URL и наименования листов
    urlSheets = fun.codesSheets(wb)
    # список url-вкладок для удаления
    url2del = ['SR_0420503_R3_5',
               'SR_0420503_R4_5',
               'SR_0420503_R4']

    for url in url2del:
        sheetName = fun.sheetNameFromUrl(urlSheets, url)
        wb.remove(wb[sheetName])

def scha_prirost(id_fond, wb_xbrl, df_avancor):
    """ Копируем данные в формы Прироста """
    # ================================================================
    def prirost_01():
        """ 0420503 Отчет о приросте об уме -> SR_0420503_R1 """
        shortURL = 'SR_0420503_R1'
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя нужной формы
        ws_xbrl = wb_xbrl[sheetName]
        print(f'{sheetName} - {shortURL}')

        row_start = ws_xbrl.max_row
        cols_start = [x for x in range (1, ws_xbrl.max_column + 1)]

        # Вкладка с которой копируем данные
        # 0420502 Справка о стоимости чис   'SR_0420502_R1'
        shortURL_from = 'SR_0420502_R1'
        sheetName_from = fun.sheetNameFromUrl(urlSheets, shortURL_from)  # имя нужной формы
        ws_xbrl_from = wb_xbrl[sheetName_from]
        row_start_from = ws_xbrl_from.max_row
        # список копирумых колонок
        # копируем всё, кроме колонки №2
        cols_from = [x for x in range (1, ws_xbrl_from.max_column + 1) if x != 2]

        # копируем ячейки
        for i in range(ws_xbrl.max_column):
            ws_xbrl.cell(row_start, cols_start[i]).value = \
                ws_xbrl_from.cell(row_start_from, cols_from[i]).value

        # Проверяем форму на наличие не заполненных ячеек
        cellBegin = 'A10'
        cellEnd = 'F10'
        dCheck.empty_cell(ws_xbrl, cellBegin, cellEnd)

    # ================================================================
    def prirost_02():
        """ 0420503 Отчет о приросте об у_2 -> SR_0420503_R2 """
        shortURL = 'SR_0420503_R2'
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя нужной формы
        ws_xbrl = wb_xbrl[sheetName]
        print(f'{sheetName} - {shortURL}')

        row = ws_xbrl.max_row
        cols = [2, 3]

        # Вкладка с которой копируем данные
        # 0420502 Справка о стоимости ч_2   'SR_0420502_R2'
        shortURL_from = 'SR_0420502_R2'
        sheetName_from = fun.sheetNameFromUrl(urlSheets, shortURL_from)  # имя нужной формы
        ws_xbrl_from = wb_xbrl[sheetName_from]
        row_start_from = ws_xbrl_from.max_row
        # список копирумых колонок
        cols_from = [2, 4]
        # ---------------------------------------------------------
        # вставляем период
        dataEnd = ws_xbrl_from.cell(row_start_from, 2).value
        dataBbegin = dataEnd[:-2] + '01'
        ws_xbrl.cell(row, 2).value = dataBbegin + ' - ' + dataEnd
        # ---------------------------------------------------------
        # копируем код валюты
        ws_xbrl.cell(row, 3).value = \
            ws_xbrl_from.cell(row_start_from, 4).value
        # ---------------------------------------------------------
        # Записываем в форму идентификатор фонда
        dcop.copy_id_fond_to_tbl(ws_xbrl, id_fond)

    # ================================================================
    def prirost_04():
        """ 0420503 Отчет о приросте об у_4 -> SR_0420503_R3 """
        shortURL = 'SR_0420503_R3'
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя нужной формы
        ws_xbrl = wb_xbrl[sheetName]
        print(f'{sheetName} - {shortURL}')

        # Считываем данные из "Раздел III..." (файл-Аванкор)
        # Раздел III. Сведения о приросте (об уменьшении) стоимости имущества, принадлежащего
        # акционерному инвестиционному фонду (составляющего паевой инвестиционный фонд)
        row_start_av = 26
        row_end_av = 97
        col_2_av = 5    # Код строки (колонка 2)
        col_3_av = 6    # Значение показателя за отчетный период (колонка 3)
        data = {}
        for row in range(row_start_av, row_end_av + 1):
            cell_2 = df_avancor.loc[row, col_2_av]
            cell_3 = df_avancor.loc[row, col_3_av]
            if str(cell_2) != 'nan' and cell_3 != 0:
                data[cell_2] = analiz_data_all(cell_3)

        # Заполняем форму
        row_start_xbrl = 7
        row_end_xbrl = 54
        col_2_xbrl = 2
        col_3_xbrl = 3
        for row in range(row_start_xbrl, row_end_xbrl + 1):
            # номер строки в xbrl
            cell = ws_xbrl.cell(row, col_2_xbrl).value
            # сравниваем номера строк в Аванкоре и xbrl
            if cell in data.keys():
                ws_xbrl.cell(row, col_3_xbrl).value = data[cell]
                # Форматируем ячейку
                ws_xbrl.cell(row, col_3_xbrl).alignment = Alignment(horizontal='right')
        # ---------------------------------------------------------
        # Записываем в форму идентификатор фонда
        dcop.copy_id_fond_to_tbl(ws_xbrl, id_fond)

    # ================================================================
    def prirost_07():
        """ 0420503 Отчет о приросте об у_7 -> SR_0420503_podpisant """
        shortURL = 'SR_0420503_podpisant'
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя нужной формы
        ws_xbrl = wb_xbrl[sheetName]
        print(f'{sheetName} - {shortURL}')

        row = ws_xbrl.max_row
        col = ws_xbrl.max_column

        # Вкладка с которой копируем данные
        # 0420502 Справка о стоимости _56   'SR_0420502_Podpisant'
        shortURL_from = 'SR_0420502_Podpisant'
        sheetName_from = fun.sheetNameFromUrl(urlSheets, shortURL_from)  # имя нужной формы
        ws_xbrl_from = wb_xbrl[sheetName_from]
        row_from = ws_xbrl_from.max_row
        col_from = ws_xbrl_from.max_column

        # копируем фио директора УК
        ws_xbrl.cell(row, col).value = \
            ws_xbrl_from.cell(row_from, col_from).value

        # Проверяем форму на наличие не заполненных ячеек
        cellBegin = 'B7'
        cellEnd = 'B7'
        dCheck.empty_cell(ws_xbrl, cellBegin, cellEnd)


    # ================================================================
    def prirost_08():
        """ 0420503 Отчет о приросте об у_8 -> SR_0420503_podpisant_spec_dep """
        shortURL = 'SR_0420503_podpisant_spec_dep'
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя нужной формы
        ws_xbrl = wb_xbrl[sheetName]
        print(f'{sheetName} - {shortURL}')

        row = 8
        cols = [1,2,3,4,5]

        # Вкладка с которой копируем данные
        # 0420502 Справка о стоимости _57   'SR_0420502_Podpisant_spec_dep'
        shortURL_from = 'SR_0420502_Podpisant_spec_dep'
        sheetName_from = fun.sheetNameFromUrl(urlSheets, shortURL_from)  # имя нужной формы
        ws_xbrl_from = wb_xbrl[sheetName_from]

        # копируем ячейки
        for i in range(len(cols)):
            ws_xbrl.cell(row, cols[i]).value = \
                ws_xbrl_from.cell(row, cols[i]).value

        # Проверяем форму на наличие не заполненных ячеек
        cellBegin = 'A8'
        cellEnd = 'E8'
        dCheck.empty_cell(ws_xbrl, cellBegin, cellEnd)

    # ================================================================
    # список кодов всех форм
    urlSheets = fun.codesSheets(wb_xbrl)

    prirost_01()
    prirost_02()
    prirost_04()
    prirost_07()
    prirost_08()

    # ================================================================


if __name__ == "__main__":
    pass
