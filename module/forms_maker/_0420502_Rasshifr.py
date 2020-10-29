""" Формирование форм СЧА - расшифровки разделов"""

# from module.functions import listSheetsName
import module.functions as fun
import module.dataCopy as dcop
import module.adjustments as adj
from openpyxl.utils import column_index_from_string  # 'B' -> 2
from module.globals import *


# Выбираем формы: расшифровки разделов
# 0420502 Справка о стоимости _14	SR_0420502_Rasshifr_Akt_P1_P1
# 0420502 Справка о стоимости _15	SR_0420502_Rasshifr_Akt_P1_P2
# 0420502 Справка о стоимости _16	SR_0420502_Rasshifr_Akt_P2_1
# 0420502 Справка о стоимости _17	SR_0420502_Rasshifr_Akt_P2_10
# 0420502 Справка о стоимости _18	SR_0420502_Rasshifr_Akt_P2_11
# 0420502 Справка о стоимости _19	SR_0420502_Rasshifr_Akt_P2_2
# 0420502 Справка о стоимости _20	SR_0420502_Rasshifr_Akt_P2_3
# 0420502 Справка о стоимости _21	SR_0420502_Rasshifr_Akt_P2_4
# 0420502 Справка о стоимости _22	SR_0420502_Rasshifr_Akt_P2_5
# 0420502 Справка о стоимости _23	SR_0420502_Rasshifr_Akt_P2_6
# 0420502 Справка о стоимости _24	SR_0420502_Rasshifr_Akt_P2_7
# 0420502 Справка о стоимости _25	SR_0420502_Rasshifr_Akt_P2_8
# 0420502 Справка о стоимости _26	SR_0420502_Rasshifr_Akt_P2_9
# 0420502 Справка о стоимости _27	SR_0420502_Rasshifr_Akt_P3_1
# 0420502 Справка о стоимости _28	SR_0420502_Rasshifr_Akt_P3_3
# 0420502 Справка о стоимости _29	SR_0420502_Rasshifr_Akt_P3_4
# 0420502 Справка о стоимости _30	SR_0420502_Rasshifr_Akt_P3_5
# 0420502 Справка о стоимости _31	SR_0420502_Rasshifr_Akt_P3_6
# 0420502 Справка о стоимости _32	SR_0420502_Rasshifr_Akt_P3_7
# 0420502 Справка о стоимости _33	SR_0420502_Rasshifr_Akt_P4_1
# 0420502 Справка о стоимости _34	SR_0420502_Rasshifr_Akt_P4_2_1
# 0420502 Справка о стоимости _35	SR_0420502_Rasshifr_Akt_P4_2_2
# 0420502 Справка о стоимости _36	SR_0420502_Rasshifr_Akt_P5_1
# 0420502 Справка о стоимости _37	SR_0420502_Rasshifr_Akt_P5_2
# 0420502 Справка о стоимости _38	SR_0420502_Rasshifr_Akt_P5_3
# 0420502 Справка о стоимости _39	SR_0420502_Rasshifr_Akt_P5_4
# 0420502 Справка о стоимости _40	SR_0420502_Rasshifr_Akt_P5_5
# 0420502 Справка о стоимости _41	SR_0420502_Rasshifr_Akt_P6_1_1
# 0420502 Справка о стоимости _42	SR_0420502_Rasshifr_Akt_P6_1_2
# 0420502 Справка о стоимости _43	SR_0420502_Rasshifr_Akt_P6_2_1
# 0420502 Справка о стоимости _44	SR_0420502_Rasshifr_Akt_P6_2_2
# 0420502 Справка о стоимости _45	SR_0420502_Rasshifr_Akt_P7_1
# 0420502 Справка о стоимости _46	SR_0420502_Rasshifr_Akt_P7_2
# 0420502 Справка о стоимости _47	SR_0420502_Rasshifr_Akt_P7_3
# 0420502 Справка о стоимости _48	SR_0420502_Rasshifr_Akt_P7_4
# 0420502 Справка о стоимости _49	SR_0420502_Rasshifr_Akt_P7_5
# 0420502 Справка о стоимости _50	SR_0420502_Rasshifr_Akt_P7_6
# 0420502 Справка о стоимости _51	SR_0420502_Rasshifr_Akt_P7_7
# 0420502 Справка о стоимости _52	SR_0420502_Rasshifr_Akt_P8_1
# 0420502 Справка о стоимости _53	SR_0420502_Rasshifr_Akt_P8_2
# 0420502 Справка о стоимости _54	SR_0420502_Rasshifr_Ob_P1
# 0420502 Справка о стоимости _55	SR_0420502_Rasshifr_Ob_P2
# 0420502 Справка о стоимости _58	SR_0420502_Rasshifr_Akt_P3_2

def correct_RegNumberKO(ws, rowBegin, col):
    """Корректируем номер кредитной организации"""
    # (поле:"Регистрационный номер кредитной организации")
    # начиная со строки 'rowBegin', колонка 'col'
    for row in range(rowBegin, ws.max_row):
        ws.cell(row, col).value = str(ws.cell(row, col).value).split('.')[0]


def rashifr(wb, df_avancor, df_identifier, id_fond):
    def copy_rashifr(ws, AvancoreTitle, copy_id=True):
        """ Копируем данные форм-расшифровок"""

        # номер первой и последней строки с данными в таблице Аванкор
        data_row, row_end = adj.data_search(df_avancor, AvancoreTitle)

        if data_row != row_end:
            # Если номера первой и последней строки НЕ совпадают, то Раздел содержит данные
            # список всех номеров строк в таблице Аванкор
            rows_numbers = [x for x in range(data_row, row_end + 1)]

            # цифра в последней колонке и номер строки с идектификаторами в таблице XBRL
            # (количество колонок для копирования)
            max_number, id_row = fun.end_col_number(ws)

            # Вставить нужное кол-во строк перед строкой "Итого"
            ws.insert_rows(ws.max_row, amount=(len(rows_numbers) - 1))

            # координаты первой ячейки в таблице XBRL
            col_begin, row_begin = fun.begin_cell(ws, max_number)

            # номера колонок с иденитификаторами в таблице XBRL
            id_cols = fun.id_nombers(col_begin)

            # Номера колонок в таблице Аванкор, за исключением пустых"""
            columns_numbers = fun.find_columns_numbers(df_avancor, collumn_max, max_number, data_row)

            # Копирование данных из таблицы Аванков в таблицу XBRL
            dcop.copy_data(ws, df_avancor, rows_numbers, columns_numbers, row_begin, col_begin)

            if copy_id:
                # Копирование идентификаторов
                # Находим и вставляем идентификаторы в ячейки
                dcop.insert_id(ws, id_row, id_cols, df_identifier, rows_numbers, row_begin, col_begin, id_fond)

            return True

        else:
            # Если номера первой и последней строки совпадают, то Раздел пуст.
            # Удаляем вкладку
            # wb.remove(ws)
            return False

    # **********************************************************************************
    def rashifr_14():
        """0420502 Справка о стоимости _14	SR_0420502_Rasshifr_Akt_P1_P1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P1_P1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'J11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '1.1. Денежные средства на счетах в кредитных организациях'

        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Корректируем номер кредитной организации
            # (поле:"Регистрационный номер кредитной организации")
            # начиная с 11 строки
            # for row in range(11, ws.max_row):
            #     ws.cell(row, 6).value = str(ws.cell(row, 6).value).split('.')[0]
            correct_RegNumberKO(ws, 11, 6)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    # **********************************************************************************
    def rashifr_14_v2():
        """0420502 Справка о стоимости _14	SR_0420502_Rasshifr_Akt_P1_P1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P1_P1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'J11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '1.1. Денежные средства на счетах в кредитных организациях'

        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Корректируем номер кредитной организации
            # (поле:"Регистрационный номер кредитной организации")
            # начиная с 11 строки
            # for row in range(11, ws.max_row):
            #     ws.cell(row, 6).value = str(ws.cell(row, 6).value).split('.')[0]
            correct_RegNumberKO(ws, 11, 6)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
            # ---------------------------------------------------------
            # копируем номер счета из столбца "Примечание"
            row_begin = 11
            col_id = 'B'
            col_from = 'L'
            # col_name = ws.max_column
            # adj.copy_cells_one2one(ws, row_begin, col_id, col_from, del_old_sell=True)
            adj.copy_bank_account(id_fond, wb_pif_info, ws, row_begin, col_id, col_from)

            # ---------------------------------------------------------
            # Вставляем идентификатор кредитной организации
            row_begin = 11
            col_id = 'C'
            col_from = 'D'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)

            # ---------------------------------------------------------

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_15():
        """0420502 Справка о стоимости _15	SR_0420502_Rasshifr_Akt_P1_P2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P1_P2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '1.2. Денежные средства на счетах по депозиту в кредитных организациях'

        # При наличии депозитов в файле-Аванкор нужно указывать номер счета депозита
        adj.depozitID(df_avancor, AvancoreTitle)

        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Корректируем номер кредитной организации
            # (поле:"Регистрационный номер кредитной организации")
            # начиная с 11 строки, колонка 6
            # for row in range(11, ws.max_row):
            #     ws.cell(row, 6).value = str(ws.cell(row, 6).value).split('.')[0]
            correct_RegNumberKO(ws, 11, 6)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_15_v2():
        """0420502 Справка о стоимости _15	SR_0420502_Rasshifr_Akt_P1_P2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P1_P2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '1.2. Денежные средства на счетах по депозиту в кредитных организациях'

        # При наличии депозитов в файле-Аванкор нужно указывать номер счета депозита
        adj.depozitID(df_avancor, AvancoreTitle)

        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Корректируем номер кредитной организации
            # (поле:"Регистрационный номер кредитной организации")
            # начиная с 11 строки, колонка 6
            # for row in range(11, ws.max_row):
            #     ws.cell(row, 6).value = str(ws.cell(row, 6).value).split('.')[0]
            correct_RegNumberKO(ws, 11, 6)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
            # ---------------------------------------------------------
            # копируем номер счета из столбца "Примечание"
            row_begin = 11
            col_id = 'B'
            col_from = 'M'
            # col_name = ws.max_column
            # adj.copy_cells_one2one(ws, row_begin, col_id, col_from, del_old_sell=True)
            adj.copy_bank_account(id_fond, wb_pif_info, ws, row_begin, col_id, col_from)

            # ---------------------------------------------------------
            # Вставляем идентификатор кредитной организации
            row_begin = 11
            col_id = 'C'
            col_from = 'D'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)

            # ---------------------------------------------------------

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_16():
        """0420502 Справка о стоимости _16	SR_0420502_Rasshifr_Akt_P2_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.1. Облигации российских хозяйственных обществ'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_17():
        """0420502 Справка о стоимости _17	SR_0420502_Rasshifr_Akt_P2_10"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_10'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.10. Ипотечные сертификаты участия'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_18():
        """0420502 Справка о стоимости _18	SR_0420502_Rasshifr_Akt_P2_11"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_11'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.11. Иные ценные бумаги российских эмитентов (за исключением закладных)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_19():
        """0420502 Справка о стоимости _19	SR_0420502_Rasshifr_Akt_P2_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.2. Государственные ценные бумаги Российской Федерации'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_20():
        """0420502 Справка о стоимости _20	SR_0420502_Rasshifr_Akt_P2_3"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_3'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.3. Государственные ценные бумаги субъектов Российской Федерации'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_21():
        """0420502 Справка о стоимости _21	SR_0420502_Rasshifr_Akt_P2_4"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_4'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.4. Муниципальные ценные бумаги'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_22():
        """0420502 Справка о стоимости _22	SR_0420502_Rasshifr_Akt_P2_5"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_5'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'S11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.5. Российские депозитарные расписки'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_23():
        """0420502 Справка о стоимости _23	SR_0420502_Rasshifr_Akt_P2_6"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_6'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.6. Инвестиционные паи паевых инвестиционных фондов'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_24():
        """0420502 Справка о стоимости _24	SR_0420502_Rasshifr_Akt_P2_7"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_7'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.7. Акции российских акционерных обществ'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

        # **********************************************************************************

    def rashifr_24_v2():
        """0420502 Справка о стоимости _24	SR_0420502_Rasshifr_Akt_P2_7"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_7'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.7. Акции российских акционерных обществ'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор Выпуска ценной бумаги
            row_begin = 11
            col_id = 'B'
            col_from = 'H'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)

            # ---------------------------------------------------------
            # Вставляем идентификатор Биржи
            row_begin = 11
            col_id = 'C'
            col_from = 'N'
            adj.copy_birzha_id(ws, row_begin, col_id, col_from)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_25():
        """0420502 Справка о стоимости _25	SR_0420502_Rasshifr_Akt_P2_8"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_8'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.8. Векселя российских хозяйственных обществ'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_26():
        """0420502 Справка о стоимости _26	SR_0420502_Rasshifr_Akt_P2_9"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P2_9'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '2.9. Облигации с ипотечным покрытием'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_27():
        """0420502 Справка о стоимости _27	SR_0420502_Rasshifr_Akt_P3_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.1. Облигации иностранных коммерческих организаций'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_28():
        """0420502 Справка о стоимости _28	SR_0420502_Rasshifr_Akt_P3_3"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_3'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.3. Облигации международных финансовых организаций'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_29():
        """0420502 Справка о стоимости _29	SR_0420502_Rasshifr_Akt_P3_4"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_4'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'S11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.4. Иностранные депозитарные расписки'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_30():
        """0420502 Справка о стоимости _30	SR_0420502_Rasshifr_Akt_P3_5"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_5'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.5. Паи (акции) иностранных инвестиционных фондов'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_31():
        """0420502 Справка о стоимости _31	SR_0420502_Rasshifr_Akt_P3_6"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_6'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.6. Акции иностранных акционерных обществ'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_32():
        """0420502 Справка о стоимости _32	SR_0420502_Rasshifr_Akt_P3_7"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_7'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'I11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.7. Иные ценные бумаги иностранных эмитентов'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_33():
        """0420502 Справка о стоимости _33	SR_0420502_Rasshifr_Akt_P4_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P4_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '4.1. Недвижимое имущество'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_33_v2():
        """0420502 Справка о стоимости _33	SR_0420502_Rasshifr_Akt_P4_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P4_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '4.1. Недвижимое имущество'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор
            row_begin = 11
            col_id = 'B'
            col_from = 'C'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_34():
        """0420502 Справка о стоимости _34	SR_0420502_Rasshifr_Akt_P4_2_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P4_2_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '4.2.1. Право аренды недвижимого имущества (арендодатель – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_35():
        """0420502 Справка о стоимости _35	SR_0420502_Rasshifr_Akt_P4_2_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P4_2_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '4.2.2. Право аренды недвижимого имущества (арендодатель – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_36():
        """0420502 Справка о стоимости _36	SR_0420502_Rasshifr_Akt_P5_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P5_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '5.1. Имущественные права из договоров участия в долевом строительстве объектов недвижимого имущества'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_37():
        """0420502 Справка о стоимости _37	SR_0420502_Rasshifr_Akt_P5_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P5_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '5.2. Имущественные права, связанные с возникновением права собственности ' \
                        'на объект недвижимости (его часть) после завершения его строительства (создание) ' \
                        'и возникающие из договора, стороной по которому является юридическое лицо, ' \
                        'которому принадлежит право собственности или иное вещное право, ' \
                        'включая право аренды, на земельный участок, выделенный в установленном порядке ' \
                        'для целей строительства объекта недвижимости, и (или) имеющее разрешение ' \
                        'на строительство объекта недвижимости на указанном земельном участке, ' \
                        'либо юридическое лицо, инвестирующее денежные средства ' \
                        'или иное имущество в строительство объекта недвижимости'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_38():
        """0420502 Справка о стоимости _38	SR_0420502_Rasshifr_Akt_P5_3"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P5_3'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '5.3. Имущественные права из договоров, на основании ' \
                        'которых осуществляется строительство (создание) объектов ' \
                        'недвижимого имущества (в том числе на месте объектов ' \
                        'недвижимости) на выделенном в установленном порядке ' \
                        'для целей строительства (создания) указанного объекта ' \
                        'недвижимости земельном участке, который (право аренды которого) ' \
                        'составляет активы акционерного инвестиционного фонда ' \
                        '(паевого инвестиционного фонда)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_39():
        """0420502 Справка о стоимости _39	SR_0420502_Rasshifr_Akt_P5_4"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P5_4'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '5.4. Имущественные права из договоров, на основании ' \
                        'которых осуществляется реконструкция объектов недвижимости, ' \
                        'составляющих активы акционерного инвестиционного фонда ' \
                        '(паевого инвестиционного фонда)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_40():
        """0420502 Справка о стоимости _40	SR_0420502_Rasshifr_Akt_P5_5"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P5_5'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'D11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '5.5. Иные имущественные права'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_41():
        """0420502 Справка о стоимости _41	SR_0420502_Rasshifr_Akt_P6_1_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_1_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.1.1. Денежные требования по кредитным договорам и договорам займа (должник – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_41_v2():
        """0420502 Справка о стоимости _41	SR_0420502_Rasshifr_Akt_P6_1_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_1_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.1.1. Денежные требования по кредитным договорам и договорам займа (должник – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор физ лица
            row_begin = 11
            col_id = 'B'
            col_from = 'J'
            col_fio = 'I'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   word_start=3, word_end=5,fio=col_fio)
            # ---------------------------------------------------------
            # Вставляем Идентификатор денежного требования
            row_begin = 11
            col_id = 'C'
            col_from = 'D'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   delta=True, dogovor_n=True, fond_name=True)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_42():
        """0420502 Справка о стоимости _42	SR_0420502_Rasshifr_Akt_P6_1_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_1_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'N11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.1.2. Денежные требования по кредитным договорам и договорам займа (должник – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_42_v2():
        """0420502 Справка о стоимости _42	SR_0420502_Rasshifr_Akt_P6_1_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_1_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'N11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.1.2. Денежные требования по кредитным договорам и договорам займа (должник – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем Идентификатор юридического лица
            row_begin = 11
            col_id = 'B'
            col_from = 'L'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)
            # ---------------------------------------------------------
            # Вставляем Идентификатор денежного требования
            row_begin = 11
            col_id = 'C'
            col_from = 'D'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   word_end=1,
                                   delta=True, dogovor_n=True, fond_name=True)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_43():
        """0420502 Справка о стоимости _43	SR_0420502_Rasshifr_Akt_P6_2_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_2_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.2.1. Закладные (должник – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_44():
        """0420502 Справка о стоимости _44	SR_0420502_Rasshifr_Akt_P6_2_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P6_2_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'N11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '6.2.2. Закладные (должник – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_45():
        """0420502 Справка о стоимости _45	SR_0420502_Rasshifr_Akt_P7_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.1. Доли в уставных капиталах российских обществ с ограниченной ответственностью'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_46():
        """0420502 Справка о стоимости _46	SR_0420502_Rasshifr_Akt_P7_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.2. Права участия в уставных капиталах иностранных коммерческих организаций'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_47():
        """0420502 Справка о стоимости _47	SR_0420502_Rasshifr_Akt_P7_3"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_3'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'H11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.3. Проектная документация для строительства или реконструкции объекта недвижимости'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_48():
        """0420502 Справка о стоимости _48	SR_0420502_Rasshifr_Akt_P7_4"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_4'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'F11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = 'Подраздел 7.4. Драгоценные металлы'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_49():
        """0420502 Справка о стоимости _49	SR_0420502_Rasshifr_Akt_P7_5"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_5'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'G11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.5. Требования к кредитной организации выплатить денежный эквивалент драгоценных металлов по текущему курсу'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_50():
        """0420502 Справка о стоимости _50	SR_0420502_Rasshifr_Akt_P7_6"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_6'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'E11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.6. Художественные ценности'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_51():
        """0420502 Справка о стоимости _51	SR_0420502_Rasshifr_Akt_P7_7"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_7'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'E11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.7. Иное имущество, не указанное в таблицах пунктов 7.1 - 7.6'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Добавляем подробное описание имущества
            adj.corrector_scha_51_(ws, df_identifier)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_51_v2():
        """0420502 Справка о стоимости _51	SR_0420502_Rasshifr_Akt_P7_7"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P7_7'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'E11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '7.7. Иное имущество, не указанное в таблицах пунктов 7.1 - 7.6'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Добавляем подробное описание имущества
            adj.corrector_scha_51_(ws, df_identifier)
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # формируем "Вид иного имущества" из столбца "Примечание"
            row_begin = 11
            col_id = 'B'
            col_from = 'G'
            # col_name = ws.max_column
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from,
                                   del_old_sell=True, id_fond=id_fond)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_52():
        """0420502 Справка о стоимости _52	SR_0420502_Rasshifr_Akt_P8_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P8_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '8.1. Дебиторская задолженность (должник – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_52_v2():
        """0420502 Справка о стоимости _52	SR_0420502_Rasshifr_Akt_P8_1"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P8_1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'K11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '8.1. Дебиторская задолженность (должник – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор физ лица
            row_begin = 11
            col_id = 'B'
            col_from = 'J'
            col_fio = 'H'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   word_start=3, word_end=5,
                                   fio=col_fio, fond_name=False)
            # ---------------------------------------------------------
            # Вставляем идентификатор основания дебиторской задолженности
            row_begin = 11
            col_id = 'C'
            col_from = 'F'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   delta=True, dogovor_n=True, fond_name=True)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_53():
        """0420502 Справка о стоимости _53	SR_0420502_Rasshifr_Akt_P8_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P8_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '8.2. Дебиторская задолженность (должник – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Корректируем ИНН
            # начиная с 11 строки, колонка 9 (I=9)
            for row in range(11, ws.max_row):
                ws.cell(row, 9).value = str(ws.cell(row, 9).value).split('.')[0]
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Убираем лишние '.00' в конце строки,
            # которые могут появиться после копирования ИНН
            adj.corrector_00(ws, row=11, col='I')
            adj.corrector_00(ws, row=11, col='J')
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_53_v2():
        """0420502 Справка о стоимости _53	SR_0420502_Rasshifr_Akt_P8_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P8_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '8.2. Дебиторская задолженность (должник – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Корректируем ИНН
            # начиная с 11 строки, колонка 9 (I=9)
            for row in range(11, ws.max_row):
                ws.cell(row, 9).value = str(ws.cell(row, 9).value).split('.')[0]
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Убираем лишние '.00' в конце строки,
            # которые могут появиться после копирования ИНН
            adj.corrector_00(ws, row=11, col='I')
            adj.corrector_00(ws, row=11, col='J')
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор дебитора
            row_begin = 11
            col_id = 'B'
            col_from = 'J'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)
            # ---------------------------------------------------------
            # Вставляем идентификатор основания задолженности
            row_begin = 11
            col_id = 'C'
            col_from = 'F'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   delta=True, dogovor_n=True, fond_name=True)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_54():
        """0420502 Справка о стоимости _54	SR_0420502_Rasshifr_Ob_P1"""

        shortURL = 'SR_0420502_Rasshifr_Ob_P1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'J11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = 'Подраздел 1. Кредиторская задолженность (кредитор – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_54_v2():
        """0420502 Справка о стоимости _54	SR_0420502_Rasshifr_Ob_P1"""

        shortURL = 'SR_0420502_Rasshifr_Ob_P1'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'J11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = 'Подраздел 1. Кредиторская задолженность (кредитор – физическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор физ лица
            row_begin = 11
            col_id = 'B'
            col_from = 'I'
            col_fio = 'H'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   word_start=3, word_end=5, fio=col_fio)
            # ---------------------------------------------------------
            # Вставляем идентификатор основания задолженности
            row_begin = 11
            col_id = 'C'
            col_from = 'F'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   delta=True, dogovor_n=True, fond_name=True)

        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_55():
        """0420502 Справка о стоимости _55	SR_0420502_Rasshifr_Ob_P2"""

        shortURL = 'SR_0420502_Rasshifr_Ob_P2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = 'Подраздел 2. Кредиторская задолженность (кредитор – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_55_v2():
        """0420502 Справка о стоимости _55	SR_0420502_Rasshifr_Ob_P2"""

        shortURL = 'SR_0420502_Rasshifr_Ob_P2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'M11'
        cellFormatN = 3

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = 'Подраздел 2. Кредиторская задолженность (кредитор – юридическое лицо)'
        if copy_rashifr(ws, AvancoreTitle, copy_id=False):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)

            # ---------------------------------------------------------
            # Вставляем идентификатор кредитора
            row_begin = 11
            col_id = 'B'
            col_from = 'I'
            adj.copy_cells_one2one(ws, row_begin, col_id, col_from)
            # ---------------------------------------------------------
            # Вставляем идентификатор основания задолженности
            row_begin = 11
            col_id = 'C'
            col_from = 'F'
            adj.copy_hash_of_cells(id_fond, ws, row_begin, col_id, col_from,
                                   word_end=1,
                                   delta=True, dogovor_n=True, fond_name=True)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    def rashifr_58():
        """0420502 Справка о стоимости _58	SR_0420502_Rasshifr_Akt_P3_2"""

        shortURL = 'SR_0420502_Rasshifr_Akt_P3_2'  # код вкладки
        sheetName = fun.sheetNameFromUrl(urlSheets, shortURL)  # имя вкладки
        ws = wb[sheetName]
        # ячейки для форматирования
        cellFormat = 'L11'
        cellFormatN = 2

        print(f'{sheetName} - {shortURL}')
        # ---------------------------------------------------------
        # Переносим данные в форму:
        # Заголовки формы в файле-Аванкор
        AvancoreTitle = '3.2. Облигации иностранных государств'
        if copy_rashifr(ws, AvancoreTitle):
            # ---------------------------------------------------------
            # Записываем в форму идентификатор фонда
            dcop.copy_id_fond_to_tbl(ws, id_fond)
            # ---------------------------------------------------------
            # Форматируем ячейки
            fun.cellFormat(ws, cellFormat, cols=cellFormatN)
        else:
            # Если ничего скопировано не было, то Раздел пуст.
            # Удаляем вкладку
            wb.remove(ws)

    # **********************************************************************************
    wb_pif_info = fun.load_pif_info(file_name=pif_info,
                                    path_2file=dir_shablon)

    urlSheets = fun.codesSheets(wb)  # словарь - "код вкладки":"имя вкладки"
    # Кол-во строк и столбцов в файле Аванкор
    index_max = df_avancor.shape[0]
    collumn_max = df_avancor.shape[1]

    rashifr_14_v2()
    rashifr_15_v2()
    rashifr_16()
    rashifr_17()
    rashifr_18()
    rashifr_19()
    rashifr_20()
    rashifr_21()
    rashifr_22()
    rashifr_23()
    rashifr_24_v2()
    rashifr_25()
    rashifr_26()
    rashifr_27()
    rashifr_28()
    rashifr_29()
    rashifr_30()
    rashifr_31()
    rashifr_32()
    rashifr_33_v2()
    rashifr_34()
    rashifr_35()
    rashifr_36()
    rashifr_37()
    rashifr_38()
    rashifr_39()
    rashifr_40()
    rashifr_41_v2()
    rashifr_42_v2()
    rashifr_43()
    rashifr_44()
    rashifr_45()
    rashifr_46()
    rashifr_47()
    rashifr_48()
    rashifr_49()
    rashifr_50()
    rashifr_51_v2()
    rashifr_52_v2()
    rashifr_53_v2()
    rashifr_54_v2()
    rashifr_55_v2()
    rashifr_58()


# %%

if __name__ == "__main__":
    pass
