import sys
import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string  # ‘B12’ -> (‘B’, 12)
from openpyxl.utils import column_index_from_string  # 'B' -> 2
# from openpyxl.utils.cell import get_column_letter  # 3 -> 'C'
# from openpyxl.utils.cell import coordinate_to_tuple  # 'D2' -> (2,4)
from openpyxl.styles import Alignment
from module.globals import *
global log
# from module.data_load import load_pif_info


# %%
def coordinate(cell):
    """Конвртер координат: 'A10' преобразуем в '10, 1' """
    col, row = coordinate_from_string(cell)
    col = column_index_from_string(col)
    return row, col


# %%
def end_col_number(ws, row_id_fond=5):
    """ Поиск номера столбца в последней колонке таблицы XBRL """
    # row_id_fond - номер строки, в которой расположенг идентификатор фонда

    max_number = row = 0
    # номер последней колонки
    for row in range(1, 5):
        cell = ws.cell(row_id_fond + row, ws.max_column)
        if cell.value and cell.value.isdigit():  # если в ячейке цифра, то заканчиваем цикл
            max_number = int(cell.value)
            break
    if not max_number or not row:
        log.error(f'Ошибка в функции: "end_col_number()"')

    return max_number, row_id_fond + row + 1  # строка с идентификаторами ниже на одну строчку


# %%
def begin_cell(ws, max_number, row_id_fond=5):
    """ Поиск координат первой ячейки в таблице XBRL """
    # max_number - цифра в ячейке правой крайней колонки

    row_begin = 0
    col_begin = ws.max_column - (max_number - 1)
    for row in range(3, 10):
        cell = ws.cell(row_id_fond + row, col_begin)
        if cell.fill.start_color.index == '00000000':
            row_begin = row_id_fond + row
            break
    if not row_begin:
        log.error(f'.......... Ошибка в функции: "begin_cell()"')

    return col_begin, row_begin


# %%
def id_nombers(col_begin):
    """ Поиск номеров колонок с иденитификаторами в таблице XBRL"""

    # номер колонкс с "Итого"
    col_itogo = 2
    # количество колонок с идентификаторами
    id_numbers = col_begin - col_itogo

    # номера колонок с идентификаторами
    id_cols = [(col_itogo + col) for col in range(id_numbers)]

    return id_cols


# %%
def razdel_name_row(df_avancor, title_1_name, index_max, title_col=2):
    """Поиск Номера строки с названием раздела в файле Аванкор"""

    for row in range(1, index_max):
        title = str(df_avancor.loc[row, title_col])
        if title == title_1_name or title[:20] == title_1_name[:20]:
            title_row = row
            return title_row

    log.error(f'Раздел отчетности: "{title_1_name}" в таблице-Аванкор не найден')
    sys.exit("Ошибка!")


# %%
def start_data_row(df_avancor, index_max, title_row, title_col=2):
    """Поиск номера первой строки с данными в отдельной таблице Аванкор"""
    # index_max - кол-во строк в файле
    # title_col - номер колонки с названием таблицы
    # title_row - номер строки с названием таблицы

    for row in range(1, index_max):
        data_i = df_avancor.loc[title_row + row, title_col]
        if str(data_i) == '1':
            data_row = title_row + row + 1
            return data_row


# %%
def end_data_row(df_avancor, index_max, data_row, title_col=2):
    """Поиск номера последней строки с данными в отдельной таблице Аванкор"""
    # index_max - кол-во строк в файле
    # data_row  - номер первой строки с данными
    # title_col - номер колонки с названием таблицы

    data = df_avancor.loc[data_row, title_col]
    if data == 'Итого' or str(data) == 'nan' or str(data).startswith('Оценочная стоимость'):
        # (' - раздел пуст')
        row_end = data_row
        return row_end
    else:
        for row in range(1, index_max):
            data_i = df_avancor.loc[data_row + row, title_col]
            if data_i == 'Итого' or str(data_i) == 'nan':
                row_end = data_row + row
                return row_end

    return log.error(f'не найдена последняя строка с данными таблице Аванкор:'
                     f'название таблицы: "{data}"'
                     f'номер первой строки "{data_row}"')


# %%
def find_columns_numbers(df_avancor, collumn_max, max_number, data_row, data_col=3):
    """ Поиск номеров колонок в таблице Аванкор, за исключением пустых"""
    # data_row, data_col_1 - координаты первой колонки с данными в Аванкор
    # max_number - номер последней колонки в таблице XBRL

    column = []
    for i in range(0, max_number):
        while str(df_avancor.loc[data_row - 1, data_col]) == 'nan' and data_col <= collumn_max:
            data_col += 1
        column.append(data_col)
        data_col += 1
    return column


# %%
def codesSheets(wb) -> dict:
    """ Словарь: URL и наименования листов """
    codes_sheets = {}
    for sheet in wb.sheetnames:
        ws_xbrl = wb[sheet]
        ws_xbrl_cell = ws_xbrl.cell(3, 1)
        codes_sheets[ws_xbrl_cell.value] = sheet

    # исключаем из списка вкладку '_dropDownSheet'
    for code in codes_sheets:
        if codes_sheets[code] == '_dropDownSheet':
            codes_sheets.pop(code)
            break

    return codes_sheets


# %%
def sheetNameFromUrl(codesSheets: dict, shortURL: str) -> str:
    """ Поиск имени вкладки по части кода формы"""

    for url in codesSheets:
        if url.endswith(shortURL):
            return codesSheets[url]

    log.error(f'В отчетном файле не найдено имя вкладки с кодом "{shortURL}"')
    sys.exit("Ошибка!")


# %%

def listSheetsName(wb, shortURLs: list) -> list:
    """ Составляем список наименований вкладок из списка коротких 'url'"""

    SheetsNames = []
    urlSheets = codesSheets(wb)
    for url in shortURLs:
        SheetsNames.append(sheetNameFromUrl(urlSheets, url))
    return SheetsNames


# %%

def cellFormat(ws, cell, cols: int = None):
    """Форматируем ячейки (выравниваем по правому краю)"""
    # 'cols' - кол-во колонок, данные в которых будут отформатированы
    # (если 'cols' не задан, то форматируются данные во всех колонках, начиная с 'cell')

    col1, row1 = coordinate_from_string(cell)
    col1 = column_index_from_string(col1)

    colEnd = col1 + cols if cols else ws.max_column + 1
    rowEnd = ws.max_row + 1
    for row in range(row1, rowEnd):
        for col in range(col1, colEnd):
            ws.cell(row, col).alignment = Alignment(horizontal='right')


# %%
def pathToFile(up=1, folder=None):
    """ Путь к файлу, расположенному в папке 'folder' """
    # up = 1  # кол-во каталогов "вверх"
    # folder = название папки: например - 'Шаблоны'

    path_to_current_file = os.path.realpath(__file__)
    path_to_current_folder = os.path.dirname(path_to_current_file)
    path_to_folder = path_to_current_folder.split('\\')
    if up > 0:
        path_to_folder = path_to_folder[:-up]
    path_to_folder.append(folder)
    path_to_folder = '/'.join(path_to_folder)

    return path_to_folder


# %%
def fileName(file_name):
    """ Имя файла без пути к нему"""
    # пример: file_name = 'h:/_Отчетность_/Конвертер/Идентификаторы.xlsx'
    return file_name.split('/')[-1]


def fileDir(file_name):
    """ Путь к файлу"""
    # пример: file_name = 'h:/_Отчетность_/Конвертер/Идентификаторы.xlsx'
    return '/'.join(file_name.split('/')[:-1]) + '/'


# %%
def fond_id_search():
    """Построение списка идентификаторов фонда из файла"""

    # Загружаем данные из файла с информацией о фондах
    wb = openpyxl.load_workbook(dir_shablon + pif_info)

    ws, table = tabl_search(wb,
                            sheet_name=pif_info_sheet,
                            tbl_name=pif_info_tbl_name)
    # Номер колонки с названием "id"
    col = col_search(ws, table, col_name='id')

    min_col, min_row, max_col, max_row = \
        openpyxl.utils.cell.range_boundaries(table.ref)

    fond_id = []
    for row in range(min_row + 1, max_row + 1):
        fond_id.append(ws.cell(row, col).value)

    return fond_id


# %%
def tabl_search(wb, sheet_name="", tbl_name=""):
    """Поиск таблицы"""

    ws = wb[sheet_name]

    table_found = None

    for tbl in ws._tables:
        if ws._tables[tbl].name == tbl_name:
            table_found = ws._tables[tbl]  # tbl
            break
    if not table_found:
        log.error(f'Таблица "{tbl_name}" не найдена.')
        log.error(f'---выполнение программы прервано---')
        sys.exit()

    # for tbl in ws._tables:
    #     print(tbl)
    #     print(" : " + tbl.displayName)
    #     print("   -  name = " + tbl.name)
    #     print("   -  type = " + (tbl.tableType if isinstance(tbl.tableType, str) else 'n/a'))
    #     print("   - range = " + tbl.ref)
    #     print("   - #cols = %d" % len(tbl.tableColumns))
    #     for col in tbl.tableColumns:
    #         print("     : " + col.name)

    return ws, table_found


def col_search(ws, table, col_name=''):
    """Поиск номера колонки с заданным именем"""

    size = table.ref  # 'C19:D27'
    min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(size)
    col_number = None

    for col in range(min_col, max_col + 1):
        cell = ws.cell(min_row, col).value
        if cell == col_name:
            col_number = col
            break
    if not col_number:
        log.error(f'Колонка:"{col_name}" в таблице:"{table.name}" не найдена.')
        log.error(f'---выполнение программы прервано---')
        sys.exit()

    return col_number


def pai_search(fond_id):
    """Точность указания паев"""

    # Загружаем данные из файла с информацией о фондах
    wb = openpyxl.load_workbook(dir_shablon + pif_info)

    ws, table = tabl_search(wb,
                            sheet_name=pif_info_sheet,
                            tbl_name=pif_info_tbl_name)

    col_id = col_search(ws, table, col_name=tbl_col_id)
    col_pai = col_search(ws, table, col_name=tbl_col_pai)

    min_col, min_row, max_col, max_row = \
        openpyxl.utils.cell.range_boundaries(table.ref)
    row_pai = None
    for row in range(min_row + 1, max_row + 1):
        cell = ws.cell(row, col_id).value
        if cell == fond_id:
            row_pai = row
            break

    if not row_pai:
        log.error(f'Идентификатор:"{fond_id}" в таблице:"{table.name}" не найден.')
        log.error(f'---выполнение программы прервано---')
        sys.exit(1)

    return ws.cell(row_pai, col_pai).value


# %%
if __name__ == "__main__":
    pass
